from datetime import datetime
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db import transaction
from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from .models import (
    Applicant, ApplicantHistory, DeletedApplicant, AcceptedApplicant, Queue, STATUS_CHOICES
)
from .forms import (
    ApplicantCreateForm, ApplicantEditFormHRHelp, ApplicantEditFormHR,
    ExperienceFormSet, AcceptedFollowUpForm
)
from .utils import is_hr, is_hr_help,is_hr_or_hr_help

from django.contrib.auth.decorators import login_required
from django.shortcuts import render

@login_required
def hr_dashboard(request):
    return render(request, "hr/hr_dashboard.html")

@login_required
def hr_help_dashboard(request):
    return render(request, "hr/hr_help_dashboard.html")


# ------------------------------------------------------------------------------
# HELPER: سجل حركة/تاريخ
# ------------------------------------------------------------------------------
def _log_history(applicant, user, action, changes=None):
    ApplicantHistory.objects.create(
        applicant=applicant,
        action=action,
        updated_by=user,
        changes=changes or ""
    )

# مقارنة بسيطة بين قيم قديمة/جديدة عشان نسجل "إيه اللي اتغير"
def _diff_changes(instance, old_data: dict, fields):
    diffs = []
    for f in fields:
        old = old_data.get(f, None)
        new = getattr(instance, f, None)
        if str(old) != str(new):
            diffs.append(f"{f}: '{old}' -> '{new}'")
    return "; ".join(diffs)

# ------------------------------------------------------------------------------
# 1) إنشاء طلب جديد (HR_HELP فقط)
# ------------------------------------------------------------------------------
@login_required
@user_passes_test(is_hr_or_hr_help)
@transaction.atomic
def applicant_create(request):
    if request.method == "POST":
        form = ApplicantCreateForm(request.POST, request.FILES)
        formset = ExperienceFormSet(request.POST, prefix="exp")
        if form.is_valid() and formset.is_valid():
            # حماية تكرار الرقم القومي برسالة أوضح
            nid = form.cleaned_data.get("national_id")
            if Applicant.objects.filter(national_id=nid).exists():
                messages.error(request, "هذا الرقم القومي مسجّل بالفعل — لا يمكن تكرار الطلب.")
                return redirect("hr:applicant_search_or_create")

            applicant = form.save(commit=False)
            applicant.created_by = request.user
            applicant.last_updated_by = request.user
            applicant.save()

            formset.instance = applicant
            formset.save()

            _log_history(applicant, request.user, "create", changes="إنشاء طلب جديد")

            messages.success(request, f"تم حفظ الطلب #{applicant.order_number} بنجاح.")
            return redirect("hr:applicant_detail", order_number=applicant.order_number)
        else:
            messages.error(request, "برجاء مراجعة المدخلات.")
    else:
        form = ApplicantCreateForm()
        formset = ExperienceFormSet(prefix="exp")

    return render(request, "hr/applicant_create.html", {
        "form": form,
        "formset": formset,
        "page_title": "تسجيل طلب جديد",
    })

# شاشة بحث سريعة قبل الإنشاء (يدخل الرقم القومي)
@login_required
@user_passes_test(is_hr_or_hr_help)
def applicant_search_or_create(request):
    nid = request.GET.get("national_id", "").strip()
    found = None
    if nid:
        found = Applicant.objects.filter(national_id=nid).first()
        if found:
            messages.warning(request, f"المتقدم برقم قومي {nid} لديه طلب سابق #{found.order_number}.")
    return render(request, "hr/applicant_search_or_create.html", {
        "found": found,
        "query_nid": nid,
    })

# ------------------------------------------------------------------------------
# 2) عرض قائمة الطلبات + فلترة (HR يرى الكل / HR_HELP يرى ما أنشأه)
# ------------------------------------------------------------------------------
@login_required
def applicant_list(request):
    qs = Applicant.objects.all().order_by("order_number")
    role_hr = is_hr(request.user)
    role_help = is_hr_help(request.user)

    if role_help and not role_hr:
        qs = qs.filter(created_by=request.user)  # الموظف العادي يشوف طلباته فقط

    # فلاتر بسيطة: من-إلى تاريخ التقديم + حالة + بحث (اسم/قومي/موبايل)
    date_from = request.GET.get("from")
    date_to = request.GET.get("to")
    status = request.GET.get("status")
    q = request.GET.get("q")

    if date_from:
        try:
            dt = datetime.strptime(date_from, "%Y-%m-%d")
            qs = qs.filter(created_at__date__gte=dt.date())
        except ValueError:
            pass
    if date_to:
        try:
            dt = datetime.strptime(date_to, "%Y-%m-%d")
            qs = qs.filter(created_at__date__lte=dt.date())
        except ValueError:
            pass
    if status and status in dict(STATUS_CHOICES):
        qs = qs.filter(status=status)
    if q:
        qs = qs.filter(
            # بحث بسيط
            # اسم أو قومي أو هاتف
            # ملاحظة: لو عايز بحث OR أكثر تفصيلاً استخدم Q
            full_name__icontains=q
        ) | qs.filter(national_id__icontains=q) | qs.filter(phone__icontains=q)

    return render(request, "hr/applicant_list.html", {
        "applicants": qs,
        "status_choices": STATUS_CHOICES,
        "filters": {"from": date_from, "to": date_to, "status": status, "q": q},
    })

# ------------------------------------------------------------------------------
# 3) تفاصيل طلب
# ------------------------------------------------------------------------------
@login_required
def applicant_detail(request, order_number):
    applicant = get_object_or_404(Applicant, order_number=order_number)
    # الموظف العادي لا يرى إلا ما أنشأه
    if is_hr_help(request.user) and not is_hr(request.user):
        if applicant.created_by_id != request.user.id:
            raise Http404("غير مسموح لك بالاطلاع على هذا الطلب.")
    return render(request, "hr/applicant_detail.html", {
        "applicant": applicant,
    })

# ------------------------------------------------------------------------------
# 4) تعديل طلب (HR_HELP يعدل طلباته، HR يعدل أي طلب)
# ------------------------------------------------------------------------------
@login_required
@transaction.atomic
def applicant_edit(request, order_number):
    applicant = get_object_or_404(Applicant, order_number=order_number)
    role_hr = is_hr(request.user)
    role_help = is_hr_help(request.user)

    # صلاحيات الوصول
    if role_help and not role_hr and applicant.created_by_id != request.user.id:
        raise Http404("غير مسموح لك بتعديل هذا الطلب.")

    # اختيار الفورم حسب الدور
    FormClass = ApplicantEditFormHR if role_hr else ApplicantEditFormHRHelp

    if request.method == "POST":
        form = FormClass(request.POST, request.FILES, instance=applicant)
        formset = ExperienceFormSet(request.POST, prefix="exp", instance=applicant)
        if form.is_valid() and formset.is_valid():
            # خزن قيم قديمة عشان نسجل الفرق
            tracked_fields = [f.name for f in Applicant._meta.fields if f.name not in ("created_at", "updated_at")]
            old_data = {f: getattr(applicant, f) for f in tracked_fields}

            obj = form.save(commit=False)
            obj.last_updated_by = request.user
            obj.save()
            formset.save()

            # سجل الاختلافات
            changes = _diff_changes(obj, old_data, tracked_fields)
            _log_history(obj, request.user, "update", changes=changes or "تعديل بيانات")

            messages.success(request, "تم حفظ التعديلات.")
            return redirect("hr:applicant_detail", order_number=order_number)
        else:
            messages.error(request, "برجاء مراجعة المدخلات.")
    else:
        form = FormClass(instance=applicant)
        formset = ExperienceFormSet(prefix="exp", instance=applicant)

    return render(request, "hr/applicant_edit.html", {
        "form": form,
        "formset": formset,
        "applicant": applicant,
    })

# ------------------------------------------------------------------------------
# 5) حذف طلب (HR فقط) + تسجيل في DeletedApplicant + History
# ------------------------------------------------------------------------------
@login_required
@user_passes_test(is_hr)
@transaction.atomic
def applicant_delete(request, order_number):
    applicant = get_object_or_404(Applicant, order_number=order_number)

    if request.method == "POST":
        # سجل نسخة للمحذوفات
        DeletedApplicant.objects.create(
            original_order_number=applicant.order_number,
            full_name=applicant.full_name,
            national_id=applicant.national_id,
            phone=applicant.phone,
            email=applicant.email,
            deleted_by=request.user,
        )
        _log_history(applicant, request.user, "delete", changes="حذف الطلب")

        applicant.delete()
        messages.success(request, "تم حذف الطلب، ويمكنك مراجعته في صفحة المحذوفات.")
        return redirect("hr:applicant_list")

    return render(request, "hr/applicant_delete_confirm.html", {
        "applicant": applicant,
    })

# ------------------------------------------------------------------------------
# 6) اتخاذ قرار HR (قبول/رفض/احتياطي) + إنشاء نسخة بالمقبولين
# ------------------------------------------------------------------------------
@login_required
@user_passes_test(is_hr)
@transaction.atomic
def applicant_decision(request, order_number):
    applicant = get_object_or_404(Applicant, order_number=order_number)
    if request.method == "POST":
        decision = request.POST.get("decision")  # accepted/rejected/reserve
        if decision not in dict(STATUS_CHOICES):
            messages.error(request, "قرار غير صالح.")
            return redirect("hr:applicant_detail", order_number=order_number)

        applicant.status = decision
        applicant.decision_by = request.user
        applicant.decision_at = timezone.now()
        applicant.last_updated_by = request.user
        applicant.save()

        # لو مقبول — أنشئ سجل AcceptedApplicant لو مش موجود
        if decision == "accepted":
            AcceptedApplicant.objects.get_or_create(applicant=applicant)

        _log_history(applicant, request.user, "decision", changes=f"قرار: {decision}")
        messages.success(request, "تم حفظ القرار.")
        return redirect("hr:applicant_detail", order_number=order_number)

    return render(request, "hr/applicant_decision.html", {
        "applicant": applicant,
        "status_choices": STATUS_CHOICES,
    })

# ------------------------------------------------------------------------------
# 7) قائمة المقبولين (فلتر تاريخ)
# ------------------------------------------------------------------------------
@login_required
def accepted_list(request):
    qs = AcceptedApplicant.objects.select_related("applicant").all().order_by("applicant__order_number")
    date_from = request.GET.get("from")
    date_to = request.GET.get("to")
    if date_from:
        try:
            dt = datetime.strptime(date_from, "%Y-%m-%d")
            qs = qs.filter(applicant__decision_at__date__gte=dt.date())
        except ValueError:
            pass
    if date_to:
        try:
            dt = datetime.strptime(date_to, "%Y-%m-%d")
            qs = qs.filter(applicant__decision_at__date__lte=dt.date())
        except ValueError:
            pass

    return render(request, "hr/accepted_list.html", {
        "accepted": qs,
        "filters": {"from": date_from, "to": date_to},
    })

# ------------------------------------------------------------------------------
# 8) قائمة المحذوفات (فلتر تاريخ)
# ------------------------------------------------------------------------------
@login_required
@user_passes_test(is_hr)
def deleted_list(request):
    qs = DeletedApplicant.objects.all().order_by("-snapshot_at")
    date_from = request.GET.get("from")
    date_to = request.GET.get("to")
    if date_from:
        try:
            dt = datetime.strptime(date_from, "%Y-%m-%d")
            qs = qs.filter(snapshot_at__date__gte=dt.date())
        except ValueError:
            pass
    if date_to:
        try:
            dt = datetime.strptime(date_to, "%Y-%m-%d")
            qs = qs.filter(snapshot_at__date__lte=dt.date())
        except ValueError:
            pass

    return render(request, "hr/deleted_list.html", {
        "deleted": qs,
        "filters": {"from": date_from, "to": date_to},
    })

# ------------------------------------------------------------------------------
# 9) شاشة الدور (HR و HR_HELP يشوفوا / HR يضغط التالي)
# ------------------------------------------------------------------------------
def _get_next_pending(current_order_number=None):
    qs = Applicant.objects.filter(status="pending").order_by("order_number")
    if current_order_number:
        return qs.filter(order_number__gt=current_order_number).first() or qs.first()
    return qs.first()

@login_required
def queue_view(request):
    queue, _ = Queue.objects.get_or_create(id=1)
    # أول مرة: لو مفيش حد محدد في الدور — حدد أول pending
    if not queue.current_applicant:
        queue.current_applicant = _get_next_pending()
        queue.save()

    # الموظف العادي يشوف بس (شاشة خضراء/تنبيه برقم الطلب الحالي)
    return render(request, "hr/queue.html", {
        "current": queue.current_applicant,
        "can_next": is_hr(request.user),  # زرار التالي يظهر لـ HR فقط
    })

@login_required
@user_passes_test(is_hr)
def queue_next(request):
    queue, _ = Queue.objects.get_or_create(id=1)
    current = queue.current_applicant
    next_app = _get_next_pending(current.order_number if current else None)
    queue.current_applicant = next_app
    queue.save()
    if next_app:
        messages.success(request, f"التالي: طلب #{next_app.order_number} - {next_app.full_name}")
    else:
        messages.info(request, "لا يوجد طلبات قيد الانتظار.")
    return redirect("hr:queue_view")

# ------------------------------------------------------------------------------
# 10) تصدير إكسل (طلب واحد / مجمّع)
# ------------------------------------------------------------------------------
import openpyxl
from openpyxl.utils import get_column_letter

def _worksheet_autofit(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                value = str(cell.value) if cell.value is not None else ""
                if len(value) > max_length:
                    max_length = len(value)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

@login_required
def export_applicant_excel(request, order_number):
    applicant = get_object_or_404(Applicant, order_number=order_number)
    # صلاحيات الاطلاع: HR يشوف الكل / HR_HELP يشوف طلباته
    if is_hr_help(request.user) and not is_hr(request.user):
        if applicant.created_by_id != request.user.id:
            raise Http404("غير مسموح لك.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"طلب #{applicant.order_number}"

    rows = [
        ("رقم الطلب", applicant.order_number),
        ("الاسم", applicant.full_name),
        ("الرقم القومي", applicant.national_id),
        ("الهاتف", applicant.phone),
        ("الحالة الاجتماعية", applicant.get_marital_status_display()),
        ("الجنسية", applicant.get_nationality_display()),
        ("النوع", applicant.get_gender_display()),
        ("الديانة", applicant.get_religion_display()),
        ("الموقف من التجنيد", applicant.get_military_status_display()),
        ("البريد", applicant.email or ""),
        ("اسم القريب", applicant.relative_name),
        ("هاتف القريب", applicant.relative_phone),
        ("مدخن", "نعم" if applicant.is_smoker else "لا"),
        ("وسيلة/سيارة", applicant.get_vehicle_ownership_display()),

        ("المؤهل", applicant.get_edu_degree_display() if applicant.edu_degree else ""),
        ("سنة التخرج", applicant.grad_year or ""),
        ("جهة الحصول", applicant.edu_institution or ""),
        ("التخصص", applicant.specialization or ""),
        ("دراسات عليا", applicant.postgrad_study or ""),
        ("التقدير", applicant.get_edu_grade_display() if applicant.edu_grade else ""),

        ("الوظيفة المتقدم لها", applicant.get_job_applied_display()),
        ("كود الوظيفة", applicant.job_code or ""),
        ("تاريخ التقديم", applicant.submitted_at.strftime("%Y-%m-%d %H:%M:%S")),
        ("سبق التقديم بالشركة", "نعم" if applicant.prev_applied else "لا"),
        ("أقارب بالشركة", "نعم" if applicant.has_relatives_in_company else "لا"),
        ("تفاصيل أقارب بالشركة", applicant.relatives_in_company or ""),
        ("أقارب/أصدقاء بشركات منافسة", "نعم" if applicant.has_relatives_in_competitors else "لا"),
        ("تفاصيل أقارب بشركات منافسة", applicant.relatives_in_competitors or ""),
        ("مشاكل صحية", "نعم" if applicant.has_health_issues else "لا"),
        ("تفاصيل المشاكل الصحية", applicant.health_issues_details or ""),

        ("الحالة الحالية", applicant.get_status_display()),
        ("أُنشئ بواسطة", applicant.created_by.get_full_name() if applicant.created_by else ""),
        ("تاريخ الإنشاء", applicant.created_at.strftime("%Y-%m-%d %H:%M:%S")),
        ("آخر تعديل بواسطة", applicant.last_updated_by.get_full_name() if applicant.last_updated_by else ""),
        ("آخر تعديل", applicant.updated_at.strftime("%Y-%m-%d %H:%M:%S")),
        ("القرار بواسطة", applicant.decision_by.get_full_name() if applicant.decision_by else ""),
        ("تاريخ القرار", applicant.decision_at.strftime("%Y-%m-%d %H:%M:%S") if applicant.decision_at else ""),
    ]
    ws.append(["الحقل", "القيمة"])
    for r in rows:
        ws.append(r)

    # شيت للخبرات السابقة
    ws2 = wb.create_sheet("الخبرات السابقة")
    ws2.append(["جهة العمل", "الوظيفة", "السنوات", "الراتب", "سبب ترك العمل"])
    for exp in applicant.experiences.all():
        ws2.append([exp.employer, exp.job_title, exp.years, float(exp.salary), exp.reason_for_leaving])

    _worksheet_autofit(ws)
    _worksheet_autofit(ws2)

    filename = f"applicant_{applicant.order_number}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@login_required
def export_applicants_excel(request):
    # HR: الكل — HR_HELP: بتاعه فقط
    qs = Applicant.objects.all().order_by("order_number")
    if is_hr_help(request.user) and not is_hr(request.user):
        qs = qs.filter(created_by=request.user)

    # فلاتر اختيارية: تاريخ من/إلى
    date_from = request.GET.get("from")
    date_to = request.GET.get("to")
    if date_from:
        try:
            dt = datetime.strptime(date_from, "%Y-%m-%d")
            qs = qs.filter(created_at__date__gte=dt.date())
        except ValueError:
            pass
    if date_to:
        try:
            dt = datetime.strptime(date_to, "%Y-%m-%d")
            qs = qs.filter(created_at__date__lte=dt.date())
        except ValueError:
            pass

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "الطلبات"
    ws.append([
        "رقم الطلب", "الاسم", "الرقم القومي", "الهاتف", "الحالة", "تاريخ التقديم", "أُنشئ بواسطة"
    ])
    for a in qs:
        ws.append([
            a.order_number, a.full_name, a.national_id, a.phone,
            a.get_status_display(),
            a.submitted_at.strftime("%Y-%m-%d %H:%M:%S"),
            a.created_by.get_full_name() if a.created_by else "",
        ])

    _worksheet_autofit(ws)
    filename = "applicants.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
