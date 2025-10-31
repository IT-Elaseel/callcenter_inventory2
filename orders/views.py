# ==============================================
# 📌 Python Standard Library
# ==============================================
from datetime import datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from urllib.parse import urlencode
# ==============================================
# 📌 Third-party Libraries
# ==============================================
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
# ==============================================
# 📌 Django Imports
# ==============================================
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import (
    authenticate, login, logout, update_session_auth_hash
)
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q, Count
from django.http import (
    HttpResponse, JsonResponse, HttpResponseForbidden
)
from django.shortcuts import (
    render, redirect, get_object_or_404
)
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
from django.utils.timezone import now, localdate
from django.views.decorators.http import require_POST
# ==============================================
# 📌 Local Application Imports
# ==============================================
from .decorators import role_required
from .forms import (
    CategoryForm, ProductForm, BranchForm,
    UserCreateForm, ArabicPasswordChangeForm
)
from .models import (
    Category, Product, Branch, SecondCategory,
    Inventory, Reservation, Customer,
    InventoryTransaction, DailyRequest,
    OrderCounter, StandardRequest,
    ProductionTemplate, ProductionRequest
)
def to_decimal_safe(value, places=2):
    """حوّل أي قيمة إلى Decimal مقنّن بعدد أماكن عشرية (افتراضي 2)."""
    try:
        d = Decimal(str(value))
        quant = Decimal('1').scaleb(-places)  # Decimal('0.01') لو places=2
        return d.quantize(quant, rounding=ROUND_HALF_UP)
    except (InvalidOperation, TypeError, ValueError):
        return Decimal('0').quantize(Decimal('1').scaleb(-places))
#-----------------------------------------------------
def unit_allows_fraction(unit: str) -> bool:
    """يسمح بالكسور لو الوحدة 'kg' فقط."""
    return (unit or "").lower() == "kg"
#-----------------------التحقق من المستخدم ادمن اول لا-------
def is_admin(user):
    return (
        user.is_superuser
        or user.groups.filter(name="admin").exists()
        or (hasattr(user, "userprofile") and user.userprofile.role == "admin")
    )
def is_control(user):
    return user.is_authenticated and user.userprofile.role == "control"
#-----تصدير الحجوزات الى اكسيل-----------------------------
def export_reservations_excel(request, branch_id):
    reservations = Reservation.objects.filter(branch_id=branch_id).select_related("product", "branch", "customer").order_by("-created_at")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reservations"

    # Header
    ws.append(["ID", "Customer", "Phone", "Product", "Branch", "Delivery Type", "Status", "Created At"])

    # Data
    for r in reservations:
        ws.append([
            r.id,
            r.customer.name if r.customer else "",
            r.customer.phone if r.customer else "",
            r.product.name if r.product else "",
            r.branch.name if r.branch else "",
            r.get_delivery_type_display(),
            r.get_status_display(),
            r.created_at.strftime("%Y-%m-%d %H:%M"),
        ])

    # Response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="reservations_branch_{branch_id}.xlsx"'
    wb.save(response)
    return response
#-------------------------------------------------------------
def broadcast_new_reservation(reservation, qty=1, user=None):
    """دالة موحدة لبث الحجز الجديد لجميع الفروع"""
    channel_layer = get_channel_layer()
    async_to_sync(channel_layer.group_send)(
        "branch_updates",
        {
            "type": "branch_update",
            "message": f"🆕 حجز جديد في فرع {reservation.branch.name} ({reservation.product.name} × {qty})",
            "reservation_id": reservation.id,
            "product_name": reservation.product.name,
            "quantity": qty,
            "customer_name": reservation.customer.name if reservation.customer else "-",
            "customer_phone": reservation.customer.phone if reservation.customer and reservation.customer.phone else "-",
            "created_at": str(reservation.created_at.strftime("%Y-%m-%d %H:%M:%S")),
            "reserved_by": user.username if user else "-",
        }
    )
#-------------------------------------------------------------
@login_required
def callcenter(request):
    query = request.GET.get("q")
    category_id = request.GET.get("category")

    inventories = Inventory.objects.select_related("product", "branch", "product__category").filter(quantity__gt=0)
    categories = Category.objects.all()

    if query:
        inventories = inventories.filter(product__name__icontains=query)
    if category_id:
        inventories = inventories.filter(product__category_id=category_id)

    if request.method == "POST":
        try:
            product_id = request.POST.get("product_id")
            branch_id = request.POST.get("branch_id")
            customer_name = (request.POST.get("customer_name") or "").strip()
            customer_phone = (request.POST.get("customer_phone") or "").strip()
            delivery_type = request.POST.get("delivery_type") or "pickup"

            # ✅ هات المنتج والفرع الأول عشان نعرف وحدة المنتج
            product = get_object_or_404(Product, id=product_id)
            branch = get_object_or_404(Branch, id=branch_id)

            # ✅ قراءة الكمية بشكل آمن
            raw_qty = (request.POST.get("quantity") or "1").strip()
            try:
                q = Decimal(str(raw_qty))
            except Exception:
                return JsonResponse({"success": False, "message": "❌ كمية غير صالحة."}, status=400)

            # ✅ تحقق حسب الوحدة
            if product.unit == "kg":
                qty = q.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                if qty <= 0:
                    return JsonResponse({"success": False, "message": "❌ الكمية بالكيلو لازم تكون أكبر من 0."}, status=400)
            else:
                # عدد/سرفيز/صاج → أعداد صحيحة فقط
                qty_int = int(q.to_integral_value(rounding=ROUND_HALF_UP))
                if qty_int < 1:
                    return JsonResponse({"success": False, "message": "❌ الكمية لازم تكون عددًا صحيحًا موجبًا."}, status=400)
                qty = Decimal(qty_int)

            # ✅ معاملة وتحديث المخزون
            with transaction.atomic():
                inventory = Inventory.objects.select_for_update().get(product=product, branch=branch)

                if inventory.quantity < qty:
                    return JsonResponse(
                        {"success": False, "message": f"❌ الكمية المطلوبة غير متوفرة (المتاح {inventory.quantity})."},
                        status=400
                    )

                customer = None
                if customer_name or customer_phone:
                    customer = Customer.objects.create(
                        name=customer_name if customer_name else "عميل مؤقت",
                        phone=customer_phone if customer_phone else ""
                    )

                reservation = Reservation.objects.create(
                    customer=customer,
                    product=product,
                    branch=branch,
                    delivery_type=delivery_type,
                    status="pending",
                    quantity=qty,
                    reserved_by=request.user,
                )

                # خصم وتثبيت بدقتين عشريتين
                inventory.quantity = (inventory.quantity - qty).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                inventory.save()

            # ✅ WebSocket: ابعت أرقام كـ string لتفادي Decimal serialization
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                "callcenter_updates",
                {
                    "type": "callcenter_update",
                    "action": "upsert",
                    "product_id": product.id,
                    "product_name": product.name,
                    "category_name": product.category.name if product.category else "",
                    "branch_id": branch.id,
                    "branch_name": branch.name,
                    "new_qty": str(inventory.quantity),  # ← مهم
                    "unit": product.get_unit_display(),
                    "message": f"📦 تم تحديث {product.name} في فرع {branch.name} إلى {inventory.quantity}",
                },
            )

            async_to_sync(channel_layer.group_send)(
                "branch_updates",
                {
                    "type": "branch_update",
                    "message": f"🆕 حجز جديد ({product.name} × {str(qty)})",
                    "reservation_id": reservation.id,
                    "product_name": product.name,
                    "quantity": str(qty),  # ← مهم
                    "customer_name": customer.name if customer else "-",
                    "customer_phone": customer.phone if customer else "-",
                    "created_at": timezone.localtime(reservation.created_at).strftime('%Y-%m-%d %H:%M:%S'),
                    "reserved_by": request.user.username,
                },
            )

            async_to_sync(channel_layer.group_send)(
                "reservations_updates",
                {
                    "type": "reservations_update",
                    "action": "new",
                    "message": f"🆕 تم إضافة حجز جديد #{reservation.id}",
                    "reservation_id": reservation.id,
                    "product_name": product.name,
                    "quantity": str(qty),  # ← مهم
                    "customer_name": customer.name if customer else "-",
                    "customer_phone": customer.phone if customer else "-",
                    "branch_name": branch.name,
                    "delivery_type": reservation.get_delivery_type_display(),
                    "status": reservation.get_status_display(),
                    "created_at": timezone.localtime(reservation.created_at).strftime('%Y-%m-%d %H:%M:%S'),
                    "decision_at": "",
                    "reserved_by": request.user.username,
                },
            )

            return JsonResponse({
                "success": True,
                "message": f"✅ تم حجز {product.name}" + (f" للعميل {customer.name}" if customer else ""),
                "new_qty": str(inventory.quantity),  # ← لتوحيد النوع
            })

        except Inventory.DoesNotExist:
            return JsonResponse({"success": False, "message": "❌ لا يوجد مخزون لهذا المنتج في الفرع المختار."}, status=400)
        except Exception as e:
            import traceback; traceback.print_exc()
            return JsonResponse({"success": False, "message": f"❌ خطأ داخلي: {str(e)}"}, status=500)

    # GET
    return render(request, "orders/callcenter.html", {
        "categories": categories,
        "inventories": inventories,
        "selected_category": int(category_id) if category_id else None,
        "query": query,
    })
#----------------------------قايمه الحجوزات------------------
@login_required
def reservations_list(request):
    from datetime import date as dt_date
    today = timezone.localdate()

    # القيم من GET
    start_raw = request.GET.get("start_date", "")
    end_raw   = request.GET.get("end_date", "")
    query     = request.GET.get("q", "").strip()

    # الافتراضي: النهارده
    if not start_raw or not end_raw:
        start_date = end_date = today
        start_raw, end_raw = today.isoformat(), today.isoformat()
    else:
        try:
            start_date = dt_date.fromisoformat(start_raw)
            end_date   = dt_date.fromisoformat(end_raw)
        except ValueError:
            messages.error(request, "⚠️ صيغة التاريخ غير صحيحة.")
            start_date = end_date = today
            start_raw, end_raw = today.isoformat(), today.isoformat()

    # فلترة الحجوزات
    profile = getattr(request.user, "userprofile", None)
    reservations = Reservation.objects.all()

    if profile and profile.role == "branch":
        branch = profile.branch
        reservations = reservations.filter(branch=branch)

    reservations = reservations.filter(created_at__date__range=[start_date, end_date])

    # 🔎 البحث باسم العميل أو رقم تليفونه
    if query:
        reservations = reservations.filter(
            Q(customer__name__icontains=query) |
            Q(customer__phone__icontains=query)
        )

    reservations = reservations.select_related(
        "product", "branch", "customer"
    ).order_by("-created_at")

    return render(
        request,
        "orders/reservations.html",
        {
            "reservations": reservations,
            "user_role": profile.role if profile else None,
            "start_date": start_raw,
            "end_date": end_raw,
            "query": query,
            "today": today,  # ← أضفها
        },
    )
#-------------------------------------------------------------
# def update_reservation_status(request, res_id, status):
#     reservation = get_object_or_404(Reservation, id=res_id)
#     profile = getattr(request.user, "userprofile", None)
#     is_admin = profile and profile.role == "admin"
#
#     # ✅ تحديث الحالة في قاعدة البيانات
#     if status == "confirmed":
#         reservation.confirm(user=request.user, is_admin=is_admin)
#         msg = f"✅ تم تأكيد الحجز للعميل {reservation.customer}"
#         messages.success(request, msg)
#     elif status == "cancelled":
#         reservation.cancel(user=request.user, is_admin=is_admin)
#         msg = f"❌ تم إلغاء الحجز للعميل {reservation.customer}"
#         messages.warning(request, msg)
#     else:
#         messages.error(request, "⚠️ حالة غير معروفة")
#         return redirect(request.META.get("HTTP_REFERER", "branch_dashboard"))
#     # 🕒 حدث توقيت آخر إجراء للفرع
#     reservation.branch_last_modified_at = timezone.now()
#     reservation.save(update_fields=["branch_last_modified_at"])
#
#     # 🔁 مهم جدًا: نرجّع نحمل نسخة حديثة من قاعدة البيانات
#     reservation.refresh_from_db()
#     # ============================================================
#     # 🔄 إرسال إشعار لتحديث صفحة الحجوزات عبر WebSocket
#     # ============================================================
#     channel_layer = get_channel_layer()
#     async_to_sync(channel_layer.group_send)(
#         "reservations_updates",
#         {
#             "type": "reservations_update",      # ← لازم يطابق اسم الدالة في consumer
#             "action": "status_change",          # نميّز نوع التحديث
#             "message": msg,
#             "reservation_id": reservation.id,
#             "customer_name": reservation.customer.name if reservation.customer else "-",
#             "customer_phone": reservation.customer.phone if reservation.customer else "-",
#             "product_name": reservation.product.name,
#             "quantity": reservation.quantity,
#             "branch_name": reservation.branch.name,
#             "delivery_type": reservation.get_delivery_type_display(),
#             "status": reservation.get_status_display(),
#             "created_at": timezone.localtime(reservation.created_at).strftime('%Y-%m-%d %H:%M:%S'),
#             "decision_at": timezone.localtime(reservation.decision_at).strftime('%Y-%m-%d %H:%M:%S') if reservation.decision_at else "",
#             "branch_last_modified_at": timezone.localtime(reservation.branch_last_modified_at).strftime('%Y-%m-%d %H:%M:%S') if reservation.branch_last_modified_at else "-",
#             "reserved_by": reservation.reserved_by.username if reservation.reserved_by else "-",
#         },
#     )
#
#     return redirect(request.META.get("HTTP_REFERER", "branch_dashboard"))
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
from django.utils import timezone
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from .models import Reservation, Inventory


def update_reservation_status(request, res_id, status):
    # 🟢 نحضر بيانات الحجز المطلوبة
    reservation = get_object_or_404(Reservation, id=res_id)
    profile = getattr(request.user, "userprofile", None)
    is_admin = profile and profile.role == "admin"

    # 🧩 جروب السوكيت اللى هنبعتله التحديث
    channel_layer = get_channel_layer()

    # -------------------------------------------------------------
    # 🟢 الحالة 1: تأكيد أو إعادة تأكيد الحجز
    # -------------------------------------------------------------
    if status == "confirmed":
        old_status = reservation.status  # نحتفظ بالحالة القديمة قبل التغيير

        # ✅ استدعاء دالة التأكيد في الموديل (بتغير الحالة وتسجل الوقت والمستخدم)
        reservation.confirm(user=request.user, is_admin=is_admin)

        msg = f"✅ تم تأكيد الحجز للعميل {reservation.customer}"
        messages.success(request, msg)

        # 🟢 لو الحالة القديمة كانت 'cancelled' → يبقى دي إعادة تأكيد → نخصم الكمية تاني
        if old_status == "cancelled":
            try:
                inv = Inventory.objects.get(product=reservation.product, branch=reservation.branch)
                inv.quantity -= reservation.quantity
                if inv.quantity < 0:
                    inv.quantity = 0
                inv.save(update_fields=["quantity"])

                # 🔄 تحديث لحظي للكول سنتر عبر WebSocket
                async_to_sync(channel_layer.group_send)(
                    "callcenter_updates",
                    {
                        "type": "callcenter_update",
                        "action": "inventory_update",
                        # "message": f"📦 تم خصم {reservation.quantity} من {reservation.product.name} في {reservation.branch.name} بعد إعادة التأكيد.",
                        "product_id": reservation.product.id,
                        "product_name": reservation.product.name,
                        "category_name": getattr(reservation.product.category, 'name', ''),
                        "branch_id": reservation.branch.id,
                        "branch_name": reservation.branch.name,
                        "new_qty": float(inv.quantity),
                        "unit": reservation.product.get_unit_display(),
                    },
                )
            except Inventory.DoesNotExist:
                print("⚠️ لا يوجد سجل مخزون لهذا المنتج في هذا الفرع")

    # -------------------------------------------------------------
    # 🟠 الحالة 2: إلغاء الحجز
    # -------------------------------------------------------------
    elif status == "cancelled":
        reservation.cancel(user=request.user, is_admin=is_admin)
        msg = f"❌ تم إلغاء الحجز للعميل {reservation.customer}"
        messages.warning(request, msg)

        # 🔁 استرجاع الكمية للمخزون
        try:
            inv = Inventory.objects.get(product=reservation.product, branch=reservation.branch)
            inv.quantity += reservation.quantity
            inv.save(update_fields=["quantity"])

            # 🔄 إشعار WebSocket للكول سنتر
            async_to_sync(channel_layer.group_send)(
                "callcenter_updates",
                {
                    "type": "callcenter_update",
                    "action": "inventory_update",
                    # "message": f"🔄 تم إرجاع {reservation.quantity} من {reservation.product.name} إلى {reservation.branch.name} بعد الإلغاء.",
                    "product_id": reservation.product.id,
                    "product_name": reservation.product.name,
                    "category_name": getattr(reservation.product.category, 'name', ''),
                    "branch_id": reservation.branch.id,
                    "branch_name": reservation.branch.name,
                    "new_qty": float(inv.quantity),
                    "unit": reservation.product.get_unit_display(),
                },
            )
        except Inventory.DoesNotExist:
            print("⚠️ لا يوجد سجل مخزون مطابق لهذا الحجز")

    # -------------------------------------------------------------
    # 🔴 الحالة 3: حالة غير معروفة
    # -------------------------------------------------------------
    else:
        messages.error(request, "⚠️ حالة غير معروفة")
        return redirect(request.META.get("HTTP_REFERER", "reservations_list"))

    # -------------------------------------------------------------
    # 🕒 تحديث توقيت آخر تعديل للفرع
    # -------------------------------------------------------------
    reservation.branch_last_modified_at = timezone.now()
    reservation.save(update_fields=["branch_last_modified_at"])
    reservation.refresh_from_db()

    # -------------------------------------------------------------
    # 📢 إشعار لحظي إلى صفحة الحجوزات (Reservations Dashboard)
    # -------------------------------------------------------------
    async_to_sync(channel_layer.group_send)(
        "reservations_updates",
        {
            "type": "reservations_update",
            "action": "status_change",
            "message": msg,
            "reservation_id": reservation.id,
            "customer_name": reservation.customer.name if reservation.customer else "-",
            "customer_phone": reservation.customer.phone if reservation.customer else "-",
            "product_name": reservation.product.name,
            "quantity": float(reservation.quantity),
            "branch_name": reservation.branch.name,
            "delivery_type": reservation.get_delivery_type_display(),
            "status": reservation.get_status_display(),
            "created_at": timezone.localtime(reservation.created_at).strftime('%Y-%m-%d %H:%M:%S'),
            "decision_at": timezone.localtime(reservation.decision_at).strftime('%Y-%m-%d %H:%M:%S') if reservation.decision_at else "",
            "branch_last_modified_at": timezone.localtime(reservation.branch_last_modified_at).strftime('%Y-%m-%d %H:%M:%S'),
            "reserved_by": reservation.reserved_by.username if reservation.reserved_by else "-",
        },
    )

    # ✅ رجوع لنفس الصفحة بعد الإجراء
    return redirect(request.META.get("HTTP_REFERER", "reservations_list"))
#-------------------------------------------------------------
@login_required
@role_required(["admin"])
def reports(request):
    from datetime import date as dt_date
    today = timezone.localdate()


    start_raw = request.GET.get("start_date", "")
    end_raw   = request.GET.get("end_date", "")

    # لو مفيش قيم جاية من الفورم → الافتراضي = اليوم
    if not start_raw or not end_raw:
        start_date = end_date = today
        start_raw, end_raw = today.isoformat(), today.isoformat()
    else:
        # نحاول نفكّر التواريخ (ونتحقق من ترتيبها)
        try:
            start_date = dt_date.fromisoformat(start_raw)
            end_date   = dt_date.fromisoformat(end_raw)
        except ValueError:
            messages.error(request, "⚠️ صيغة التاريخ غير صحيحة.")
            return render(
                request,
                "orders/reports.html",
                {
                    "stats": {"total": 0, "confirmed": 0, "pending": 0, "cancelled": 0},
                    "top_products": [],
                    "top_branches": [],
                    "start_date": start_raw,
                    "end_date": end_raw,
                },
            )

        if start_date > end_date:
            messages.error(request, "⚠️ تاريخ البداية لا يجوز أن يكون بعد تاريخ النهاية.")
            return render(
                request,
                "orders/reports.html",
                {
                    "stats": {"total": 0, "confirmed": 0, "pending": 0, "cancelled": 0},
                    "top_products": [],
                    "top_branches": [],
                    "start_date": start_raw,
                    "end_date": end_raw,
                },
            )
    # لو وصلنا هنا يبقى عندنا start_date/end_date صالحين
    reservations = Reservation.objects.filter(created_at__date__range=[start_date, end_date])

    stats = {
        "total": reservations.count(),
        "confirmed": reservations.filter(status="confirmed").count(),
        "pending": reservations.filter(status="pending").count(),
        "cancelled": reservations.filter(status="cancelled").count(),
    }
    top_products = (
        reservations.values("product__name").annotate(total=Count("id")).order_by("-total")[:5]
    )

    top_branches = (
        reservations.values("branch__name").annotate(total=Count("id")).order_by("-total")[:5]
    )

    return render(
        request,
        "orders/reports.html",
        {
            "stats": stats,
            "top_products": top_products,
            "top_branches": top_branches,
            "start_date": start_raw,  # نبعث القيم كـ string عشان input يفضل ثابت
            "end_date": end_raw,
        },
    )
#-------------------------------------------------------------
@login_required
@role_required(["admin", "branch"])
def export_reports_excel(request):
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    reservations = Reservation.objects.all()
    if start_date:
        reservations = reservations.filter(created_at__gte=start_date)
    if end_date:
        reservations = reservations.filter(created_at__lte=end_date)

    # ✅ الإحصائيات
    stats = {
        "total": reservations.count(),
        "confirmed": reservations.filter(status="confirmed").count(),
        "pending": reservations.filter(status="pending").count(),
        "cancelled": reservations.filter(status="cancelled").count(),
    }

    # ✅ أكثر المنتجات
    top_products = (
        reservations.values("product__name")
        .annotate(total=Count("id"))
        .order_by("-total")[:10]
    )

    # ✅ أفضل الفروع
    top_branches = (
        reservations.values("branch__name")
        .annotate(total=Count("id"))
        .order_by("-total")[:10]
    )

    # ✅ إنشاء ملف Excel
    wb = openpyxl.Workbook()

    # Sheet 1: Stats
    ws1 = wb.active
    ws1.title = "Stats"
    ws1.append(["إحصائية", "القيمة"])
    ws1.append(["إجمالي الحجوزات", stats["total"]])
    ws1.append(["✅ Confirmed", stats["confirmed"]])
    ws1.append(["🕒 Pending", stats["pending"]])
    ws1.append(["❌ Cancelled", stats["cancelled"]])

    # Sheet 2: Top Products
    ws2 = wb.create_sheet("Top Products")
    ws2.append(["Product", "Total Reservations"])
    for p in top_products:
        ws2.append([p["product__name"], p["total"]])

    # Sheet 3: Top Branches
    ws3 = wb.create_sheet("Top Branches")
    ws3.append(["Branch", "Total Reservations"])
    for b in top_branches:
        ws3.append([b["branch__name"], b["total"]])

    # ✅ Response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="reports.xlsx"'
    wb.save(response)
    return response
#-------------------------------------------------------------
@login_required
@role_required(["branch", "admin", "callcenter"])
def branch_dashboard(request):
    profile = getattr(request.user, "userprofile", None)

    # 👇 لسsو المستخدم كول سنتر → يروح على الكول سنتر داشبورد
    if profile and profile.role == "callcenter":
        return redirect("callcenter_dashboard")

    # لو Admin → يقدر يختار أي فرع من Dropdown
    if request.user.is_superuser or (profile and profile.role == "admin"):
        selected_branch_id = request.GET.get("branch")
        branches = Branch.objects.all()

        if selected_branch_id:
            branch = Branch.objects.get(id=selected_branch_id)
            inventories = Inventory.objects.filter(branch=branch).select_related(
                "product", "product__category"
            )
            reservations = Reservation.objects.filter(branch=branch).select_related("product")
        else:
            branch = None
            inventories = Inventory.objects.none()
            reservations = Reservation.objects.none()

        return render(
            request,
            "orders/branch_dashboard.html",
            {
                "branch": branch,
                "inventories": inventories,
                "reservations": reservations,
                "is_admin": True,
                "branches": branches,
                "selected_branch_id": selected_branch_id,
            },
        )

    # لو موظف فرع عادي
    branch = profile.branch if profile else None
    if not branch:
        return render(request, "orders/branch_no_access.html")

    # تحديث الكمية (لموظف الفرع فقط)
    if request.method == "POST":
        inv_id = request.POST.get("inventory_id")
        qty = request.POST.get("quantity")
        if inv_id and qty is not None:
            inv = Inventory.objects.get(id=inv_id, branch=branch)
            inv.quantity = int(qty)
            inv.save()

    inventories = Inventory.objects.filter(branch=branch).select_related(
        "product", "product__category"
    )
    reservations = Reservation.objects.filter(branch=branch).select_related("product")

    return render(
        request,
        "orders/branch_dashboard.html",
        {
            "branch": branch,
            "inventories": inventories,
            "reservations": reservations,
            "is_admin": False,
        },
    )
#--------------------------------------------------------------
def login_view(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            # Redirect based on role
            profile = getattr(user, "userprofile", None)
            if profile:
                if profile.role == "admin":
                    return redirect("reports")
                elif profile.role == "callcenter":
                    return redirect("callcenter_dashboard")
                elif profile.role == "branch":
                    return redirect("branch_dashboard")
                elif profile.role == "control":   # ✅ جديد
                    return redirect("control_requests")
            return redirect("home")
        else:
            return render(request, "orders/login.html", {"error": "❌ بيانات الدخول غير صحيحة"})
    return render(request, "orders/login.html")
#--------------------------------------------------------------
def logout_view(request):
    logout(request)
    return redirect("login")
#--------------------------------------------------------------
def root_redirect(request):
    if not request.user.is_authenticated:
        return redirect("login")

    profile = getattr(request.user, "userprofile", None)

    if profile:
        if profile.role == "admin":
            return redirect("reports")
        elif profile.role == "callcenter":
            return redirect("callcenter")
        elif profile.role == "branch":
            return redirect("branch_dashboard")
        elif profile.role == "control":   # ✅ جديد
            return redirect("control_requests")
        elif profile.role == "hr":
            return redirect("hr:hr_dashboard")   # 👈 لسه نعملها في الـ hr app
        elif profile.role == "hr_help":
            return redirect("hr:hr_help_dashboard")  # 👈 برضو في hr app
        elif profile.role == "production":
            return redirect("production_overview")

    # fallback لو مفيش role
    return redirect("login")
#--------------------------------------------------------------
@login_required
@role_required(["branch", "admin"])
def export_inventory_excel(request, branch_id=None):
    profile = getattr(request.user, "userprofile", None)

    # 🎯 لو Admin أو Superuser
    if request.user.is_superuser or (profile and profile.role == "admin"):
        if not branch_id:
            return HttpResponse("🚫 لازم تحدد فرع في الرابط", status=400)
        try:
            branch = Branch.objects.get(id=branch_id)
        except Branch.DoesNotExist:
            return HttpResponse("🚫 الفرع المطلوب غير موجود", status=404)

    # 🎯 لو موظف فرع
    elif profile and profile.role == "branch":
        branch = profile.branch
        if not branch:
            return HttpResponse("🚫 لا يوجد فرع مربوط بحسابك", status=400)

    else:
        return HttpResponse("🚫 غير مصرح لك", status=403)

    # ✅ جلب البيانات
    inventories = Inventory.objects.filter(branch=branch).select_related("product", "product__category")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{branch.name} Inventory"

    # Header
    headers = ["Product", "Category", "Quantity", "Price"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    # Data rows
    for inv in inventories:
        ws.append([
            inv.product.name,
            inv.product.category.name if inv.product.category else "",
            inv.quantity,
            inv.product.price,
        ])

    # Auto-fit columns + borders
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
            cell.alignment = alignment
            cell.border = thin_border
        ws.column_dimensions[column].width = max_length + 2

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{branch.name}_inventory.xlsx"'
    wb.save(response)
    return response
#--------------------------------------------------------------
@role_required(["admin", "callcenter"])
def customers_list(request):
    query = request.GET.get("q")
    customers = Customer.objects.all()

    if query:
        customers = customers.filter(name__icontains=query) | customers.filter(phone__icontains=query)

    # ✅ Pagination: 10 عملاء في كل صفحة
    paginator = Paginator(customers, 10)
    page_number = request.GET.get("page")
    customers_page = paginator.get_page(page_number)

    return render(request, "orders/customers_list.html", {
        "customers": customers_page,
    })
#---------------------------------------------------------------
def landing(request):
    error_message = None

    if request.method == "POST":
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect("root_redirect")
        else:
            error_message = "❌ اسم المستخدم أو كلمة المرور غير صحيحة."
    else:
        form = AuthenticationForm()

    return render(request, "orders/landing.html", {"form": form, "login_error": error_message})
#---------------------------------------------------------------
def _get_worklist(request):
    """
    ترجع dict بالشكل: {product_id(str): qty_str}
    حيث qty_str محفوظ كسلسلة منسقة مثل "1.50"
    """
    wl = request.session.get("inventory_worklist", {})
    # نرجع نسخة نظيفة (نتجاهل القيم الغير صالحة)
    clean = {}
    for k, v in (wl or {}).items():
        try:
            pid = str(int(k))
            # نتأكد إن القيمة قابلة للتحويل لـ Decimal
            d = to_decimal_safe(v, places=2)
            if d >= Decimal('0.00'):
                clean[pid] = str(d.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
        except Exception:
            continue
    # احفظ النسخة المنظفة في الجلسة (optional)
    request.session["inventory_worklist"] = clean
    request.session.modified = True
    return clean
#---------------------------------------------------------------
def _save_worklist(request, wl_dict):
    """
    يتوقع wl_dict شكل: {pid: qty_str_or_number}
    يقوم بتخزين قيم صالحة كسلاسل منسقة في session.
    """
    safe_dict = {}
    for k, v in (wl_dict or {}).items():
        try:
            pid = str(int(k))
        except Exception:
            continue
        try:
            d = to_decimal_safe(v, places=2)
            if d >= Decimal('0.00'):
                safe_dict[pid] = str(d.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
        except Exception:
            continue
    request.session["inventory_worklist"] = safe_dict
    request.session.modified = True
#---------------------------------------------------------------
@login_required
@role_required(["branch"])
def update_inventory(request):
    profile = getattr(request.user, "userprofile", None)
    branch = profile.branch if profile else None

    if not branch:
        return render(
            request,
            "orders/no_permission.html",
            {
                "error_message": "🚫 لا يوجد فرع مربوط بحسابك. من فضلك تواصل مع مدير النظام لو محتاج صلاحية"
            },
            status=403
        )

    # القائمة المؤقتة (اللي هانشتغل عليها)
    worklist = _get_worklist(request)

    # الاستامبا (للعرض فقط لما تضغط تحميل)
    stamp_items = None

    # 🟢 POST = عمليات على القائمة/التطبيق على المخزون
    if request.method == "POST":

        # ✅ تحميل استامبا "تحديث المخزون" ودمجها في القائمة المؤقتة (من غير ما نعدل الاستامبا نفسها)
        if "load_stamp" in request.POST:
            stamp_qs = StandardRequest.objects.filter(
                branch=branch,
                stamp_type="inventory"
            ).select_related("product")

            if not stamp_qs.exists():
                messages.warning(request, "⚠️ لا توجد استامبا لتحديث المخزون لهذا الفرع.")
            else:
                updated = 0
                for it in stamp_qs:
                    pid = str(it.product_id)
                    new_val = str(to_decimal_safe(it.default_quantity, places=2))
                    # 🟢 في الحالة الجديدة بنحدث الكمية حتى لو المنتج موجود مسبقًا
                    if pid in worklist:
                        # فقط لو القيمة مختلفة فعلاً، نحدّثها
                        if worklist[pid] != new_val:
                            worklist[pid] = new_val
                            updated += 1
                    else:
                        # منتج جديد مش في القائمة
                        worklist[pid] = new_val
                        updated += 1

                _save_worklist(request, worklist)
                messages.success(request, f"✅ تم تحميل استامبا تحديث المخزون وتحديث {updated} منتج بالقيم الجديدة.")
            stamp_items = stamp_qs
        # ➕ إضافة منتج من الشبكة السفلية إلى القائمة
        elif "add_item" in request.POST:
            product_id = request.POST.get("product")
            qty = request.POST.get("quantity", "1")
            try:
                pid = str(int(product_id))
                q = to_decimal_safe(qty, places=2)
                if q < 0:
                    q = Decimal('0.00')

                # احصل على القيمة الحالية من worklist (محفوظة كسلسلة) وحولها لـ Decimal
                existing = to_decimal_safe(worklist.get(pid, '0'), places=2)
                new_total = (existing + q).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                worklist[pid] = str(new_total)
                _save_worklist(request, worklist)
                pr_name = Product.objects.get(id=int(pid)).name
                messages.success(request, f"✅ تم إضافة {pr_name} ({q}).")
            except Exception:
                messages.error(request, "❌ بيانات غير صحيحة للإضافة.")
            return redirect("update_inventory")

        # ✏️ تعديل كمية عنصر واحد داخل الجدول
        elif "update_item" in request.POST:
            rid = request.POST.get("request_id")  # هنا هي product_id
            new_qty = request.POST.get("new_quantity")
            try:
                pid = str(int(rid))
                q = to_decimal_safe(new_qty, places=2)
                if q < 0:
                    q = Decimal('0.00')
                if pid in worklist:
                    worklist[pid] = str(q.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
                    _save_worklist(request, worklist)
                    messages.success(request, "✅ تم تحديث الكمية.")
            except Exception:
                messages.error(request, "❌ لم يتم تحديث الكمية.")
            return redirect("update_inventory")

        # 🗑️ حذف عنصر واحد
        elif "delete_item" in request.POST:
            rid = request.POST.get("request_id")  # هنا هي product_id
            try:
                pid = str(int(rid))
                if pid in worklist:
                    worklist.pop(pid, None)
                    _save_worklist(request, worklist)
                    messages.success(request, "🗑️ تم حذف المنتج من القائمة.")
            except Exception:
                pass
            return redirect("update_inventory")

        # 🗑️ حذف المحدد
        elif "delete_selected" in request.POST:
            selected_ids = request.POST.getlist("selected_items")
            removed = 0
            for sid in selected_ids:
                pid = str(sid)
                if pid in worklist:
                    worklist.pop(pid, None)
                    removed += 1
            _save_worklist(request, worklist)
            messages.success(request, f"🗑️ تم حذف {removed} منتج/منتجات من القائمة.")
            return redirect("update_inventory")

        # 🗑️ حذف الكل
        elif "delete_all" in request.POST:
            worklist.clear()
            _save_worklist(request, worklist)
            messages.success(request, "🗑️ تم تفريغ القائمة بالكامل.")
            return redirect("update_inventory")

        # 💾 تحديث الكل (تطبيق القائمة المؤقتة على جدول Inventory فقط)
        elif "update_stamp" in request.POST:
            # لو المستخدم عدّل القيم في الجدول قبل الضغط تحديث الكل، التقطها
            # الفورم بيبعتها بالشكل quantities[PRODUCT_ID]
            for key, val in request.POST.items():
                if key.startswith("quantities[") and key.endswith("]"):
                    try:
                        pid = key[len("quantities["):-1]
                        q = to_decimal_safe(val, places=2)
                        if q < 0:
                            q = Decimal('0.00')
                        if pid in worklist:
                            worklist[pid] = str(q.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
                    except Exception:
                        continue
            _save_worklist(request, worklist)

            updated = 0
            for pid, qty_str in worklist.items():
                try:
                    product = Product.objects.get(id=int(pid))
                    inv, _ = Inventory.objects.get_or_create(branch=branch, product=product)
                    # احفظ الكمية كسِمَة Decimal (فُرْصى: Inventory.quantity يجب أن يكون DecimalField)
                    inv.quantity = to_decimal_safe(qty_str, places=2)
                    inv.save()

                    # 🔔 إشعار لحظي (أرسل الكمية كسلسلة للحفاظ على الدقة)
                    channel_layer = get_channel_layer()
                    async_to_sync(channel_layer.group_send)(
                        "callcenter_updates",
                        {
                            "type": "callcenter_update",
                            "action": "upsert",
                            "product_id": product.id,
                            "product_name": product.name,
                            "category_name": product.category.name if product.category else "",
                            "branch_id": branch.id,
                            "branch_name": branch.name,
                            "new_qty": str(inv.quantity),
                            "unit": product.get_unit_display(),
                            # "message": f"📦 تم تحديث {product.name} في فرع {branch.name} إلى {inv.quantity}",
                        }
                    )

                    # سجل حركة - اختيار نوع الحقل في InventoryTransaction.quantity للتحويل
                    txn_qty_decimal = to_decimal_safe(qty_str, places=2)
                    field_type = InventoryTransaction._meta.get_field('quantity').get_internal_type()
                    if field_type == 'DecimalField':
                        txn_value = txn_qty_decimal
                    else:
                        # fallback: لو لسه IntegerField → نقربه لأقرب عدد صحيح (يمكن تغييره إلى floor/ceil حسب رغبتك)
                        txn_value = int(txn_qty_decimal.to_integral_value(rounding=ROUND_HALF_UP))

                    InventoryTransaction.objects.create(
                        product=product,
                        from_branch=None,
                        to_branch=branch,
                        quantity=txn_value,
                        transaction_type="transfer_in",
                        added_by=request.user
                    )
                    updated += 1
                except Exception:
                    # لو فيه أي خطأ في صف معين نتجاهله ونكمّل
                    continue

            messages.success(request, f"✅ تم تحديث الكميات لعدد {updated} منتج.")
            return redirect("update_inventory")

        else:
            return JsonResponse({"success": False, "message": "❌ طلب غير معروف"})

    # 🟢 GET = عرض الصفحة
    categories = Category.objects.all()
    selected_category = request.GET.get("category")

    if selected_category == "":
        request.session["selected_category"] = None
        selected_category = None
    elif selected_category is not None:
        request.session["selected_category"] = selected_category
    else:
        selected_category = request.session.get("selected_category")

    products = Product.objects.filter(is_available=True)
    if selected_category:
        products = products.filter(category_id=selected_category)

    inventories = Inventory.objects.filter(branch=branch).select_related("product")
    second_categories = SecondCategory.objects.all()


    # جهّز العناصر المعروضة في الجدول (من worklist)
    work_items = []
    if worklist:
        # رجّع كل المنتجات مرة واحدة
        plist = Product.objects.filter(id__in=[int(k) for k in worklist.keys()]).select_related("category")
        prod_map = {str(p.id): p for p in plist}

        for pid, qty_str in worklist.items():
            p = prod_map.get(str(pid))
            if not p:
                continue

            # 👇 التحويل المضمون
            try:
                display_qty = Decimal(str(qty_str)).quantize(Decimal('0.01'))
            except Exception:
                display_qty = Decimal('0.00')

            # 👇 لو المنتج بالكيلو نعرضها كما هي (مثلاً 1.25)
            # لو بالعدد نحولها لصحيح
            if p.unit != "kg":
                display_qty = display_qty.to_integral_value()

            work_items.append({
                "product": p,
                "quantity": display_qty,
            })

    return render(
        request,
        "orders/update_inventory.html",
        {
            "categories": categories,
            "second_categories": second_categories,
            "selected_category": int(selected_category) if selected_category else None,
            "products": products,
            "inventories": inventories,
            "branch": branch,
            "stamp_items": stamp_items,
            "work_items": work_items,
        },
    )
#-----------------------------inventory_branch.html----------------------------------
@login_required
@role_required(["branch", "admin", "callcenter", "production"])
def update_inventory_quantity(request):
    if request.method == "POST" and request.headers.get("x-requested-with") == "XMLHttpRequest":
        try:
            product_id = int(request.POST.get("product_id"))
            new_qty = Decimal(request.POST.get("new_quantity", "0"))
            profile = getattr(request.user, "userprofile", None)
            branch = profile.branch if profile else None

            if not branch:
                return JsonResponse({"success": False, "message": "❌ لا يوجد فرع مرتبط بالمستخدم."})

            # ✅ حدّث الكمية الفعلية في جدول Inventory
            product = Product.objects.get(id=product_id)
            inv, _ = Inventory.objects.get_or_create(branch=branch, product=product)
            inv.quantity = new_qty.quantize(Decimal("0.01"))
            inv.save()
            from channels.layers import get_channel_layer
            from asgiref.sync import async_to_sync
            # ... داخل update_inventory_quantity بعد ما تحفظ التغيير
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                "callcenter_updates",
                {
                    "type": "callcenter_update",
                    "action": "upsert",
                    "product_id": product.id,
                    "product_name": product.name,
                    "category_name": product.category.name if product.category else "",
                    "branch_id": inv.branch.id,
                    "branch_name": inv.branch.name,
                    "new_qty": str(inv.quantity),
                    "unit": product.get_unit_display(),
                    # "message": f"📦 تم تحديث {product.name} في فرع {inv.branch.name} إلى {inv.quantity}",
                },
            )
            return JsonResponse({
                "success": True,
                "message": "✅ تم تحديث الكمية بنجاح.",
                "new_qty": str(inv.quantity)
            })

        except Exception as e:
            print("❌ خطأ أثناء تحديث الكمية:", e)
            return JsonResponse({"success": False, "message": "❌ حدث خطأ أثناء التحديث."})

    return JsonResponse({"success": False, "message": "❌ طلب غير صالح."})
#---------------------------------------------------------------
@login_required
@role_required(["branch"])
def set_inventory_stamp(request):
    profile = getattr(request.user, "userprofile", None)
    branch = profile.branch if profile else None

    if not branch:
        return render(
            request,
            "orders/no_permission.html",
            {"error_message": "🚫 لا يوجد فرع مربوط بحسابك."},
            status=403
        )

    selected_category = request.session.get("selected_category")

    # 🟢 POST
    if request.method == "POST":
        # ➕ إضافة منتج
        if "add_item" in request.POST:
            product_id = request.POST.get("product")
            qty = to_decimal_safe(request.POST.get("quantity") or 0, places=2)

            if product_id:
                product = Product.objects.get(id=product_id)
                StandardRequest.objects.update_or_create(
                    branch=branch,
                    product=product,
                    stamp_type="inventory",
                    defaults={
                        "default_quantity": qty,
                        "updated_at": timezone.now(),
                    }
                )
                # ✅ لو الطلب AJAX نرجع رسالة JSON
                if request.headers.get("x-requested-with") == "XMLHttpRequest":
                    return JsonResponse({
                        "success": True,
                        "message": f"✅ تمت إضافة {product.name} بكمية {qty} {product.get_unit_display()}."
                    })
                # لو مش AJAX (مش هيحصل هنا عادة)
                messages.success(request, f"✅ تمت إضافة {product.name} بكمية {qty}.")
            return redirect("set_inventory_stamp")

        # ✏️ تعديل كمية منتج
        # elif "update_item" in request.POST:
        #     std_id = request.POST.get("update_item")
        #     new_qty = request.POST.get(f"quantities[{std_id}]")
        #     if std_id and new_qty is not None:
        #         try:
        #             sr = StandardRequest.objects.get(id=std_id, branch=branch, stamp_type="inventory")
        #             sr.default_quantity = to_decimal_safe(new_qty, places=2)
        #             sr.save()
        #             messages.success(request, f"✏️ تم تحديث {sr.product.name} إلى {sr.default_quantity}.")
        #         except StandardRequest.DoesNotExist:
        #             pass
        #     return redirect("set_inventory_stamp")
        # ✏️ تعديل كمية منتج
        elif "update_item" in request.POST:
            std_id = request.POST.get("update_item")
            new_qty = request.POST.get(f"quantities[{std_id}]")

            if std_id:
                try:
                    sr = StandardRequest.objects.get(id=std_id, branch=branch, stamp_type="inventory")

                    # ✅ تحويل الكمية بأمان (حتى لو كانت صفر أو فاضية)
                    from decimal import Decimal, InvalidOperation
                    try:
                        qty_decimal = Decimal(str(new_qty).strip())
                    except (InvalidOperation, TypeError, ValueError):
                        qty_decimal = Decimal('0.00')

                    # ✅ تأكد أنها مش سالبة (اختياري)
                    if qty_decimal < Decimal('0.00'):
                        messages.error(request, "🚫 الكمية لا يمكن أن تكون سالبة.")
                    else:
                        sr.default_quantity = qty_decimal
                        sr.save()
                        messages.success(request, f"✏️ تم تحديث {sr.product.name} إلى {sr.default_quantity}.")

                except StandardRequest.DoesNotExist:
                    messages.error(request, "⚠️ لم يتم العثور على هذا المنتج.")

            return redirect("set_inventory_stamp")

        # 🗑️ حذف منتج واحد
        elif "delete_item" in request.POST:
            std_id = request.POST.get("delete_item")
            if std_id:
                StandardRequest.objects.filter(id=std_id, branch=branch, stamp_type="inventory").delete()
                # ✅ لو الطلب جاي AJAX
                if request.headers.get("x-requested-with") == "XMLHttpRequest":
                    return JsonResponse({"success": True, "deleted_id": std_id})
                else:
                    messages.success(request, "🗑️ تم حذف المنتج بنجاح.")
            return redirect("set_inventory_stamp")

        # 🗑️ حذف المحدد فقط
        elif "delete_selected" in request.POST:
            selected_ids = request.POST.getlist("selected_items")
            if selected_ids:
                StandardRequest.objects.filter(id__in=selected_ids, branch=branch, stamp_type="inventory").delete()
                if request.headers.get("x-requested-with") == "XMLHttpRequest":
                    return JsonResponse({"success": True, "deleted_ids": selected_ids})
                else:
                    messages.success(request, "🗑️ تم حذف المنتجات المحددة بنجاح.")
            else:
                if request.headers.get("x-requested-with") == "XMLHttpRequest":
                    return JsonResponse({"success": False, "message": "⚠️ لم يتم تحديد أي عنصر."})
                else:
                    messages.warning(request, "⚠️ لم يتم تحديد أي عنصر.")
            return redirect("set_inventory_stamp")

        # 🗑️ حذف الكل
        elif "delete_all" in request.POST:
            StandardRequest.objects.filter(branch=branch, stamp_type="inventory").delete()
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({"success": True, "delete_all": True})
            else:
                messages.success(request, "🗑️ تم حذف جميع المنتجات من استامبا تحديث المخزون.")
            return redirect("set_inventory_stamp")

    # 🧩 البيانات
    products = Product.objects.filter(is_available=True)
    categories = Category.objects.all()
    second_categories = SecondCategory.objects.all()
    inventory_stamps = StandardRequest.objects.filter(
        branch=branch,
        stamp_type="inventory"
    ).select_related("product__category").order_by("product__category__name", "product__name")

    return render(request, "orders/set_inventory_stamp.html", {
        "products": products,
        "categories": categories,
        "second_categories": second_categories,
        "requests_today": inventory_stamps,  # نفس الاسم عشان الـ HTML يشتغل زي الطلبية
        "selected_category": selected_category,
        "page_title": "استامبا تحديث المخزون"
    })
#----------------------------------------------------------------
@login_required
@role_required(["branch", "admin"])
def inventory_transactions(request):
    from datetime import date as dt_date

    profile = getattr(request.user, "userprofile", None)
    branch = profile.branch if profile else None

    # استلام باراميترز الفلترة
    start_raw = request.GET.get("start_date", "")
    end_raw   = request.GET.get("end_date", "")
    category_id = request.GET.get("category")
    query = request.GET.get("q", "").strip()
    branch_filter = request.GET.get("branch")  # للأدمن

    today = timezone.localdate()

    # الافتراضي: النهارده
    if not start_raw or not end_raw:
        start_date = end_date = today
        start_raw, end_raw = today.isoformat(), today.isoformat()
    else:
        try:
            start_date = dt_date.fromisoformat(start_raw)
            end_date   = dt_date.fromisoformat(end_raw)
        # 🛑 تأمين: ماينفعش تاريخ النهاية يعدي النهاردة
            if end_date > today:
                end_date = today
                end_raw = today.isoformat()

        except ValueError:
            start_date = end_date = today
            start_raw, end_raw = today.isoformat(), today.isoformat()
            messages.error(request, "⚠️ صيغة التاريخ غير صحيحة.")


    # 🟢 لو المستخدم أدمن
    if request.user.is_superuser or request.user.groups.filter(name="admin").exists():
        transactions = InventoryTransaction.objects.filter(
            transaction_type="transfer_in"
        ).select_related("product", "added_by", "to_branch").order_by("-created_at")

        # فلترة بالفرع (لو متبعتش فرع = كل الفروع)
        if branch_filter and branch_filter != "all":
            transactions = transactions.filter(to_branch_id=branch_filter)

        branches = Branch.objects.all()

    else:
        # 🟢 موظف فرع
        if not branch:
            return render(request, "orders/branch_no_access.html")

        transactions = InventoryTransaction.objects.filter(
            to_branch=branch, transaction_type="transfer_in"
        ).select_related("product", "added_by").order_by("-created_at")

        branches = None  # الفرع ثابت

    # فلترة بالتاريخ
    transactions = transactions.filter(created_at__date__range=[start_date, end_date])

    # فلترة بالقسم
    if category_id:
        transactions = transactions.filter(product__category_id=category_id)

    # فلترة بالبحث (اسم المنتج)
    if query:
        transactions = transactions.filter(product__name__icontains=query)

    categories = Category.objects.all()

    return render(
        request,
        "orders/inventory_transactions.html",
        {
            "transactions": transactions,
            "branch": branch,
            "branches": branches,   # للأدمن فقط
            "categories": categories,
            "selected_category": int(category_id) if category_id else None,
            "selected_branch": int(branch_filter) if branch_filter and branch_filter != "all" else None,
            "start_date": start_raw,
            "end_date": end_raw,
            "query": query,
            "today": today,   # ✅ علشان نستخدمه في max
        },
    )
#-----------------------------------------------------------------
@login_required
@role_required(["branch", "admin", "callcenter","production"])
def branch_inventory(request):
    profile = getattr(request.user, "userprofile", None)
    branch = profile.branch if profile else None
    role = profile.role if profile else None

    # ⬅️ لو المستخدم مالوش فرع وكان دوره فرع (بس) → امنعه
    if role == "branch" and not branch:
        return render(request, "orders/branch_no_access.html")

    # ⬅️ جلب باراميتر القسم والبحث والفرع من GET
    category_id = request.GET.get("category")
    query = request.GET.get("q", "").strip()
    branch_filter = request.GET.get("branch")

    # 🟢 لو المستخدم أدمن أو كول سنتر أو سوبر يوزر → يشوف كل الفروع
    if (
        request.user.is_superuser
        or role == "admin"
        or role == "callcenter"
        or role == "production"
    ):
        inventories = Inventory.objects.select_related("product", "product__category", "branch")
        if branch_filter and branch_filter != "all":
            inventories = inventories.filter(branch_id=branch_filter)

        branches = Branch.objects.all()
        branch_context = None

    else:
        # 🟢 موظف فرع → يجيب فرعه فقط
        inventories = Inventory.objects.filter(branch=branch).select_related(
            "product", "product__category"
        )
        branches = None
        branch_context = branch

    # ⬅️ فلترة بالقسم
    if category_id:
        inventories = inventories.filter(product__category_id=category_id)

    # ⬅️ فلترة بالبحث
    if query:
        inventories = inventories.filter(product__name__icontains=query)

    # ✅ إخفاء المنتجات ذات الكمية صفر
    inventories = inventories.filter(quantity__gt=0)

    # ✅ نحسب الأقسام مرة من كل المخزون (قبل فلترة القسم الحالي)
    # علشان تفضل القائمة كاملة حتى بعد اختيار قسم
    all_inventories = Inventory.objects.all()

    # لو المستخدم فرع → فلترها بالفرع فقط
    if not (request.user.is_superuser or role in ["admin", "callcenter", "production"]):
        all_inventories = all_inventories.filter(branch=branch)

    # استبعد المنتجات اللي كميتها صفر
    all_inventories = all_inventories.filter(quantity__gt=0)

    # الأقسام اللي فيها منتجات متوفرة
    categories = Category.objects.filter(products__inventory__in=all_inventories).distinct().order_by("name")

    return render(
        request,
        "orders/branch_inventory.html",
        {
            "branch": branch_context,
            "inventories": inventories,
            "categories": categories,
            "branches": branches,
            "selected_category": int(category_id) if category_id else None,
            "selected_branch": int(branch_filter) if branch_filter and branch_filter != "all" else None,
            "query": query,
        },
    )
#-----------------------------------------------------------------
@login_required
@role_required(["callcenter", "admin"])
def use_customer(request, customer_id):
    """
    دالة لاختيار عميل موجود (مثلاً لو فيه تضارب أرقام)
    """
    try:
        customer = Customer.objects.get(id=customer_id)
    except Customer.DoesNotExist:
        messages.error(request, "❌ العميل غير موجود.")
        return redirect("customers_list")

    # ممكن نخزن الـ id في session مؤقتًا عشان نستخدمه في الحجز القادم
    request.session["selected_customer_id"] = customer.id
    messages.success(request, f"✅ تم اختيار العميل {customer.name} ({customer.phone})")

    return redirect("customers_list")
#-----------------------------------------------------------------
@login_required
@role_required(["callcenter", "admin"])
def add_customer(request):
    if request.method == "POST":
        name = request.POST.get("name")
        phone = request.POST.get("phone")
        address = request.POST.get("address", "")

        customer = Customer.objects.create(name=name, phone=phone, address=address)
        messages.success(request, f"✅ تم إضافة العميل {customer.name} ({customer.phone})")
        return redirect("customers_list")

    return render(request, "orders/add_customer.html")
#-----------------------------------------------------------------
@login_required
@role_required(["callcenter"])
def resolve_conflict(request):
    if request.method == "POST":
        action = request.POST.get("action")
        phone = request.POST.get("phone")
        name = request.POST.get("name")
        product_id = request.POST.get("product_id")
        branch_id = request.POST.get("branch_id")
        delivery_type = request.POST.get("delivery_type")
        qty = int(request.POST.get("quantity", 1))

        product = Product.objects.get(id=product_id)
        branch = Branch.objects.get(id=branch_id)

        # لو العميل موجود
        customer = None
        if action == "use_old":
            customer = Customer.objects.filter(phone=phone).first()

        elif action == "new_customer":
            customer = Customer.objects.create(name=name, phone=phone)

        if customer:
            try:
                inventory = Inventory.objects.get(product=product, branch=branch)
                if inventory.quantity >= qty:
                    Reservation.objects.create(
                        customer=customer,
                        product=product,
                        branch=branch,
                        delivery_type=delivery_type,
                        status="pending",
                         quantity=qty,
                        reserved_by=request.user if request.user.is_authenticated else None,
                    )
                    inventory.quantity -= qty
                    inventory.save()
                    channel_layer = get_channel_layer()
                    async_to_sync(channel_layer.group_send)(
                        "callcenter_updates",
                        {
                            "type": "callcenter_update",
                            "product_id": product.id,
                            "branch_id": branch.id,
                            "branch_name": branch.name,
                            "new_qty": inventory.quantity,
                            "message": f"📦 تم تحديث {product.name} في فرع {branch.name} إلى {inventory.quantity}",
                        }
                    )

                    messages.success(
                        request,
                        f"✅ تم حجز {qty} {product.get_unit_display()} من {product.name} للعميل {customer.name}"
                    )
                else:
                    messages.error(request, f"❌ الكمية غير متوفرة من {product.name} في فرع {branch.name}")
            except Exception as e:
                messages.error(request, f"حدث خطأ أثناء الحجز: {str(e)}")

        return redirect("callcenter_dashboard")

    return redirect("callcenter_dashboard")
#-----------------------------------------------------------------
@login_required
@user_passes_test(is_admin)
def add_user_view(request):
    if request.method == "POST":
        form = UserCreateForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("add_user")  # يرجع لنفس الصفحة بعد الحفظ
    else:
        form = UserCreateForm()
    return render(request, "orders/add_user.html", {"form": form})
#-------------------------------------------------------------------
# @login_required
# def change_password(request):
#     form = ArabicPasswordChangeForm(user=request.user, data=request.POST or None)
#     success_message = None
#     show_modal = False

#     if request.method == "POST":
#         show_modal = True  # 👈 افتح المودال دايمًا بعد POST
#         if form.is_valid():
#             user = form.save()
#             update_session_auth_hash(request, user)
#             success_message = "✅ تم تغيير كلمة المرور بنجاح."
#             form = ArabicPasswordChangeForm(user=request.user)  # reset للفورم بعد النجاح

#         return render(request, "orders/home.html", {   # غير reports.html لصفحتك
#             "password_form": form,
#             "success_message": success_message,
#             "show_modal": show_modal,
#         })

#     return redirect("home")
@login_required
def change_password(request):
    form = ArabicPasswordChangeForm(user=request.user, data=request.POST or None)
    success_message = None
    show_modal = False

    if request.method == "POST":
        show_modal = True  # علشان بعد POST المودال يفتح تلقائيًا
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            success_message = "✅ تم تغيير كلمة المرور بنجاح."
            form = ArabicPasswordChangeForm(user=request.user)  # reset للفورم بعد النجاح
        else:
            success_message = "❌ حدث خطأ أثناء تغيير كلمة المرور. تأكد من صحة البيانات."

        # ✅ خليها ترندر نفس صفحة الواجهة الأساسية اللي فيها المودال
        return render(request, "orders/base.html", {
            "password_form": form,
            "success_message": success_message,
            "show_modal": show_modal,
        })

    # ❌ بدل redirect("home") خلينا نرجّعه لنفس الصفحة الرئيسية برضو
    return render(request, "orders/base.html", {
        "password_form": form,
        "success_message": None,
        "show_modal": False,
    })
#-------------------------------------------------------------------
@login_required
@user_passes_test(is_admin)
def manage_data(request):
    categories = Category.objects.all()
    products = Product.objects.all()
    branches = Branch.objects.all()

    success_message = None

    # ✅ تعريف الفورمات
    cat_form = CategoryForm(prefix="cat")
    prod_form = ProductForm()  # 🔸 بدون prefix علشان البيانات تتربط صح
    branch_form = BranchForm(prefix="branch")

    if request.method == "POST":
        # 🔹 إضافة قسم
        if "add_category" in request.POST:
            cat_form = CategoryForm(request.POST, prefix="cat")
            if cat_form.is_valid():
                cat_form.save()
                success_message = "✅ تم إضافة القسم بنجاح"
                cat_form = CategoryForm(prefix="cat")  # reset بعد الحفظ

        # 🔹 إضافة منتج
        elif "add_product" in request.POST:
            prod_form = ProductForm(request.POST)
            if prod_form.is_valid():
                prod_form.save()
                success_message = "✅ تم إضافة المنتج بنجاح"
                prod_form = ProductForm()  # ✅ تفريغ الفورم بعد الحفظ
            else:
                # 🧩 لو في أخطاء خفية هتظهر في التيرمنال
                print("❌ أخطاء الفورم:", prod_form.errors)

        # 🔹 إضافة فرع
        elif "add_branch" in request.POST:
            branch_form = BranchForm(request.POST, prefix="branch")
            if branch_form.is_valid():
                branch_form.save()
                success_message = "✅ تم إضافة الفرع بنجاح"
                branch_form = BranchForm(prefix="branch")

    return render(request, "orders/manage_data.html", {
        "cat_form": cat_form,
        "prod_form": prod_form,
        "branch_form": branch_form,
        "categories": categories,
        "products": products,
        "branches": branches,
        "success_message": success_message,
    })
#------------------------------------------------------------------
@login_required
@user_passes_test(is_admin)
def manage_users(request):
    users = User.objects.all()

    # ✅ فلترة بالاسم
    username = request.GET.get("username", "")
    if username:
        users = users.filter(username__icontains=username)

    # ✅ فلترة بالنوع (role)
    role = request.GET.get("role", "")
    if role:
        users = users.filter(userprofile__role=role)

    # ✅ POST (حذف أو إعادة تعيين باسورد)
    if request.method == "POST":
        if "delete_user" in request.POST:
            user_id = request.POST.get("delete_user")
            User.objects.filter(id=user_id).delete()

        elif "reset_password" in request.POST:
            user_id = request.POST.get("reset_password")
            u = User.objects.get(id=user_id)
            u.set_password(settings.DEFAULT_USER_PASSWORD)
            u.save()

            if hasattr(u, "userprofile"):
                u.userprofile.last_password_reset = timezone.now()
                u.userprofile.save()

        # ✅ بعد أي أكشن: رجع لنفس الرابط بالفلترة الحالية
        return redirect(request.get_full_path())

    return render(request, "orders/manage_users.html", {
        "users": users,
        "username": username,
        "role": role,
    })
#---------------------------------------------------------------
@login_required
@user_passes_test(is_admin)
def edit_category(request, pk):
    category = get_object_or_404(Category, pk=pk)
    success = False

    # ✅ نحتفظ بالـ query string (عشان نرجع بنفس الفلاتر)
    query_params = request.GET.dict()
    query_string = f"?{urlencode(query_params)}" if query_params else ""

    if request.method == "POST":
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            # ✅ بعد الحفظ رجع لنفس الصفحة مع الفلاتر
            return redirect(reverse("view_data") + query_string)
    else:
        form = CategoryForm(instance=category)

    return render(request, "orders/edit_item.html", {
        "form": form,
        "title": "✏️ تعديل قسم",
        "success": success,
        "redirect_url": reverse("view_data") + query_string,
    })
#------------------------------------------------------------------
@login_required
@user_passes_test(is_admin)
def edit_product(request, pk):
    product = get_object_or_404(Product, pk=pk)
    success = False

    # ✅ نحتفظ بالـ query string
    query_params = request.GET.dict()
    query_string = f"?{urlencode(query_params)}" if query_params else ""

    if request.method == "POST":
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            form.save()
            # ✅ بعد الحفظ رجع لنفس الصفحة مع الفلاتر
            return redirect(reverse("view_data") + query_string)
    else:
        form = ProductForm(instance=product)

    return render(request, "orders/edit_item.html", {
        "form": form,
        "title": "✏️ تعديل منتج",
        "success": success,
        "redirect_url": reverse("view_data") + query_string,
    })
#------------------------------------------------------------------
@login_required
@user_passes_test(is_admin)
def edit_branch(request, pk):
    branch = get_object_or_404(Branch, pk=pk)
    success = False

    query_params = request.GET.dict()
    query_string = f"?{urlencode(query_params)}" if query_params else ""

    if request.method == "POST":
        form = BranchForm(request.POST, instance=branch)
        if form.is_valid():
            form.save()
            # ✅ استخدم reverse
            return redirect(reverse("view_data") + query_string)
    else:
        form = BranchForm(instance=branch)

    return render(request, "orders/edit_item.html", {
        "form": form,
        "title": "✏️ تعديل فرع",
        "success": success,
        "redirect_url": reverse("view_data") + query_string,
    })
#------------------------------------------------------------------
@login_required
@user_passes_test(is_admin)
def view_data(request):
    selected_table = request.GET.get("table", "categories")
    query = request.GET.get("q", "")
    selected_category = request.GET.get("category", "")        # القسم الرئيسي
    selected_subcategory = request.GET.get("subcategory", "")  # القسم الفرعي
    availability = request.GET.get("availability", "available")  # ✅ فلتر التوفر
    success_message = None

    # ✅ حذف العناصر
    if request.method == "POST":
        if "delete_category" in request.POST:
            Category.objects.filter(id=request.POST.get("delete_category")).delete()
        elif "delete_product" in request.POST:
            Product.objects.filter(id=request.POST.get("delete_product")).delete()
        elif "delete_branch" in request.POST:
            Branch.objects.filter(id=request.POST.get("delete_branch")).delete()

    # ✅ البيانات الأساسية
    categories = Category.objects.all()
    second_categories = SecondCategory.objects.all()
    branches = Branch.objects.all()

    # ✅ المنتجات — نبدأ بالكل ثم نفلتر حسب التوفر
    products = Product.objects.all().select_related("category", "second_category")

    # 🔽 فلترة حسب التوفر
    if availability == "available":
        products = products.filter(is_available=True)
    elif availability == "unavailable":
        products = products.filter(is_available=False)
    # else → الكل

    # 🔽 فلترة المنتجات حسب البحث والأقسام
    if selected_table == "products":
        if query:
            products = products.filter(name__icontains=query)
        if selected_category:
            products = products.filter(category_id=selected_category)
        if selected_subcategory:
            products = products.filter(second_category_id=selected_subcategory)

    # ✅ تمرير البيانات للقالب
    return render(request, "orders/view_data.html", {
        "categories": categories,
        "second_categories": second_categories,
        "branches": branches,
        "products": products,
        "selected_table": selected_table,
        "query": query,
        "selected_category": selected_category,
        "selected_subcategory": selected_subcategory,
        "availability": availability,   # ✅ مهم عشان نستخدمه في HTML
        "success_message": success_message,
    })
#-----------------------------------------------------------------
@require_POST
@login_required
@user_passes_test(is_admin)
def toggle_product_availability(request, pk):
    """تبديل حالة التوفر لمنتج"""
    try:
        product = Product.objects.get(pk=pk)
        product.is_available = not product.is_available
        product.save()
        return JsonResponse({"success": True, "new_status": product.is_available})
    except Product.DoesNotExist:
        return JsonResponse({"success": False, "error": "المنتج غير موجود"})
#-----------------------------------------------------
def get_subcategories(request):
    main_id = request.GET.get("main_id")
    subcategories = SecondCategory.objects.filter(main_category_id=main_id).values("id", "name")
    return JsonResponse(list(subcategories), safe=False)
#------------------------------------------------------
@login_required
def add_daily_request(request):
    profile2 = getattr(request.user, "userprofile", None)

    if not profile2 or profile2.role not in ["branch"]:
        return render(
            request,
            "orders/no_permission.html",
            {"error_message": "🚫 غير مسموح لك بدخول هذه الصفحة."},
            status=403
        )

    branch = profile2.branch

    order_number = request.session.get("current_order_number")
    if not order_number:
        counter, _ = OrderCounter.objects.get_or_create(id=1)
        counter.current_number += 1
        counter.save()
        order_number = str(counter.current_number)
        request.session["current_order_number"] = order_number

    selected_category = request.session.get("selected_category")

    # 🟢 POST actions
    if request.method == "POST":
        # 🔹 تحميل الطلبية القياسية
        # if "load_standard" in request.POST:
        #     standard_items = StandardRequest.objects.filter(branch=branch, stamp_type="order").select_related("product", "product__category")
        #     added = 0
        #     for item in standard_items:
        #         _, created = DailyRequest.objects.get_or_create(
        #             branch=branch,
        #             product=item.product,
        #             category=item.product.category,
        #             order_number=order_number,
        #             is_confirmed=False,
        #             defaults={
        #                 "quantity": item.default_quantity,
        #                 "created_by": request.user,
        #             }
        #         )
        #         if created:
        #             added += 1
        #     messages.success(request, f"✅ تم تحميل الطلبية القياسية لهذا الفرع (أُضيف {added}).")
        #     return redirect("add_daily_request")
        # 🔹 تحميل الطلبية القياسية
        if "load_standard" in request.POST:
            selected_stamp = request.POST.get("stamp_name") or "الاستمبا الأساسية"
            standard_items = StandardRequest.objects.filter(branch=branch, stamp_type="order", stamp_name=selected_stamp).select_related("product", "product__category")

            added = 0
            for item in standard_items:
                _, created = DailyRequest.objects.get_or_create(
                    branch=branch,
                    product=item.product,
                    category=item.product.category,
                    order_number=order_number,
                    is_confirmed=False,
                    defaults={
                        "quantity": item.default_quantity,
                        "created_by": request.user,
                    }
                )
                if created:
                    added += 1
            messages.success(request, f"✅ تم تحميل {selected_stamp} ({added} منتج).")
            return redirect("add_daily_request")

        # ➕ إضافة منتج
        elif "add_item" in request.POST:
            product_id = request.POST.get("product")
            raw_qty = (request.POST.get("quantity") or "").strip()

            # نجيب المنتج وناخد منه القسم
            try:
                product = Product.objects.get(id=product_id)
                category_id = product.category_id
            except Product.DoesNotExist:
                messages.error(request, "❌ المنتج غير موجود.")
                return redirect("add_daily_request")

            # قراءة الكمية كـ Decimal
            try:
                qty = Decimal(str(raw_qty if raw_qty != "" else "0")).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            except Exception:
                qty = Decimal('0.00')

            # منع الكميات السالبة/الصفر
            if qty <= 0:
                messages.error(request, "❌ أدخل كمية صحيحة.")
                return redirect("add_daily_request")

            # لو المنتج مش بالكيلو → نحولها لعدد صحيح
            if product.unit != "kg":
                qty = qty.to_integral_value(rounding=ROUND_HALF_UP)

            try:
                dr = DailyRequest.objects.get(
                    branch=branch,
                    category_id=category_id,
                    product_id=product_id,
                    order_number=order_number,
                    is_confirmed=False
                )
                dr.quantity = (Decimal(str(dr.quantity)) + qty).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                # عدد فقط لو مش كيلو
                if product.unit != "kg":
                    dr.quantity = dr.quantity.to_integral_value(rounding=ROUND_HALF_UP)
                dr.save()
            except DailyRequest.DoesNotExist:
                DailyRequest.objects.create(
                    branch=branch,
                    category_id=category_id,
                    product_id=product_id,
                    quantity=qty,
                    created_by=request.user,
                    order_number=order_number,
                    is_confirmed=False
                )

            request.session["selected_category"] = category_id
            return redirect("add_daily_request")

        # ✏️ تحديث كمية عنصر واحد
        elif "update_item" in request.POST:
            req_id = request.POST.get("request_id")
            new_qty_raw = (request.POST.get("new_quantity") or "").strip()
            if req_id and new_qty_raw != "":
                try:
                    dr = DailyRequest.objects.select_related("product").get(
                        id=req_id, branch=branch, order_number=order_number, is_confirmed=False
                    )

                    q = Decimal(str(new_qty_raw)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                    if q <= 0:
                        messages.error(request, "❌ الكمية يجب أن تكون أكبر من صفر.")
                        return redirect("add_daily_request")

                    if dr.product.unit != "kg":
                        q = q.to_integral_value(rounding=ROUND_HALF_UP)

                    dr.quantity = q
                    dr.save()
                except DailyRequest.DoesNotExist:
                    pass
                except Exception:
                    messages.error(request, "❌ كمية غير صالحة.")
            return redirect("add_daily_request")

        # 🗑️ حذف عنصر واحد
        elif "delete_item" in request.POST:
            req_id = request.POST.get("request_id")
            if req_id:
                DailyRequest.objects.filter(
                    id=req_id, branch=branch, order_number=order_number, is_confirmed=False
                ).delete()
            return redirect("add_daily_request")

        # 🔹 حذف المحدد
        elif "delete_selected" in request.POST:
            selected_ids = request.POST.getlist("selected_items")
            if selected_ids:
                DailyRequest.objects.filter(
                    id__in=selected_ids, branch=branch, order_number=order_number, is_confirmed=False
                ).delete()
                messages.success(request, f"🗑️ تم حذف {len(selected_ids)} عنصر.")
            else:
                messages.warning(request, "⚠️ لم يتم تحديد أي عنصر.")
            return redirect("add_daily_request")

        # 🔹 حذف الكل
        elif "delete_all" in request.POST:
            DailyRequest.objects.filter(
                branch=branch, order_number=order_number, is_confirmed=False
            ).delete()
            messages.success(request, "🚮 تم حذف جميع العناصر من الطلبية الحالية.")
            return redirect("add_daily_request")

        # ✅ تأكيد الطلبية
        elif "confirm_order" in request.POST:
            now = timezone.now()
            DailyRequest.objects.filter(
                order_number=order_number, branch=branch
            ).update(is_confirmed=True, confirmed_at=now)

            layer = get_channel_layer()
            async_to_sync(layer.group_send)(
                "control_updates",
                {
                    "type": "control_update",
                    "action": "new",
                    "message": f"🆕 طلبية جديدة من فرع {branch.name}",
                    "order_number": order_number,
                }
            )
            request.session["current_order_number"] = None
            request.session["selected_category"] = None
            return redirect("add_daily_request")

    # 🧩 البيانات
    products = Product.objects.filter(is_available=True)
    categories = Category.objects.all()
    second_categories = SecondCategory.objects.all()
    requests_today = DailyRequest.objects.filter(
        order_number=order_number, branch=branch, is_confirmed=False
    ).select_related("product__category").order_by("product__category__name", "product__name")
    # 🧩 كل أسماء الاستمبات المتاحة للفرع
    all_stamps = StandardRequest.objects.filter(
        branch=branch, stamp_type="order"
    ).values_list("stamp_name", flat=True).distinct()

    return render(request, "orders/add_daily_request.html", {
        "products": products,
        "categories": categories,
        "second_categories": second_categories,
        "requests_today": requests_today,
        "order_number": order_number,
        "selected_category": selected_category,
        "all_stamps": all_stamps,  # 🆕 هنا

    })
#------------------------------------------------------
@login_required
@role_required(["branch"])
def set_standard_request(request):
    profile = getattr(request.user, "userprofile", None)
    branch = profile.branch if profile else None

    if not branch:
        return render(
            request,
            "orders/no_permission.html",
            {"error_message": "🚫 لا يوجد فرع مربوط بحسابك."},
            status=403
        )

    selected_category = request.session.get("selected_category")

    # ⬅ اسم الاستمبا الحالية من السيشن أو الافتراضي
    current_stamp = request.session.get("current_stamp_name", "الاستمبا الأساسية")

    # 🧠 لو المستخدم اختار استمبا مختلفة من الـ select أو كتب اسم جديد
    if request.method == "POST" and "select_stamp" in request.POST:
        # chosen_name = (request.POST.get("stamp_name") or "").strip()
        chosen_name = (request.POST.get("new_stamp_name") or request.POST.get("stamp_name") or "").strip()


        # لو كتب اسم جديد فاضي؟ رجّع الافتراضي
        if chosen_name == "":
            chosen_name = "الاستمبا الأساسية"

        # خزّنه في السيشن
        request.session["current_stamp_name"] = chosen_name
        messages.info(request, f"🔄 تم اختيار/إنشاء الاستمبا: {chosen_name}")
        return redirect("set_standard_request")

    # 🗑 حذف استمبا بالكامل
    if request.method == "POST" and "delete_stamp" in request.POST:
        stamp_to_delete = request.POST.get("stamp_to_delete")  # جاية من الفورم
        if stamp_to_delete:
            deleted_count, _ = StandardRequest.objects.filter(
                branch=branch,
                stamp_type="order",
                stamp_name=stamp_to_delete
            ).delete()

            messages.success(
                request,
                f"🗑️ تم حذف الاستمبا '{stamp_to_delete}' بالكامل ({deleted_count} صنف)."
            )

            # لو المستخدم كان واقف على نفس الاستمبا اللي اتمسحت -> رجعه على الافتراضي
            if request.session.get("current_stamp_name") == stamp_to_delete:
                request.session["current_stamp_name"] = "الاستمبا الأساسية"

        else:
            messages.error(request, "⚠️ اختر استمبا قبل الحذف.")
        return redirect("set_standard_request")

    # ➕ إضافة منتج جديد
    if request.method == "POST" and "add_item" in request.POST:
        product_id = request.POST.get("product")
        qty_raw = request.POST.get("quantity", "1")

        try:
            qty = Decimal(str(qty_raw)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        except Exception:
            qty = Decimal('1.00')

        if product_id and qty > 0:
            product = Product.objects.get(id=product_id)
            StandardRequest.objects.update_or_create(
                branch=branch,
                product=product,
                stamp_type="order",
                stamp_name=current_stamp,
                defaults={
                    "default_quantity": qty,
                    "updated_at": timezone.now()
                }
            )
            return JsonResponse({"success": True, "message": f"✅ تمت إضافة {product.name} بكمية {qty} {product.get_unit_display()}."})
        else:
            return JsonResponse({"success": False, "message": "⚠️ لم يتم تحديد المنتج أو الكمية."})


    # ✏️ تحديث كمية منتج
    if request.method == "POST" and "update_item" in request.POST:
        std_id = request.POST.get("request_id") or request.POST.get("update_item")
        new_qty_raw = request.POST.get(f"new_quantity_{std_id}") or request.POST.get("new_quantity")

        try:
            new_qty = Decimal(str(new_qty_raw)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        except Exception:
            new_qty = Decimal('1.00')

        if std_id and new_qty > 0:
            try:
                sr = StandardRequest.objects.get(
                    id=std_id,
                    branch=branch,
                    stamp_type="order",
                    stamp_name=current_stamp
                )
                sr.default_quantity = new_qty
                sr.save()
                messages.success(
                    request,
                    f"✏️ تم تعديل {sr.product.name} إلى {new_qty} {sr.product.get_unit_display()}."
                )
            except StandardRequest.DoesNotExist:
                messages.error(request, "❌ لم يتم العثور على العنصر.")
        return redirect("set_standard_request")

    # 🗑️ حذف المحدد
    if request.method == "POST" and "delete_selected" in request.POST:
        selected_ids = request.POST.getlist("selected_items")
        if selected_ids:
            StandardRequest.objects.filter(
                id__in=selected_ids,
                branch=branch,
                stamp_type="order",
                stamp_name=current_stamp
            ).delete()

            # ✅ لو الطلب AJAX
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({"success": True, "deleted_ids": selected_ids})
            else:
                messages.success(request, "🗑️ تم حذف العناصر المحددة.")
        else:
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({"success": False, "message": "⚠️ لم يتم تحديد أي عنصر."})
            else:
                messages.warning(request, "⚠️ لم يتم تحديد أي عنصر.")
        return redirect("set_standard_request")

    # 🗑️ حذف منتج واحد
    if request.method == "POST" and "delete_item" in request.POST or "delete_one" in request.POST:
        item_id = request.POST.get("delete_item") or request.POST.get("delete_one")
        try:
            StandardRequest.objects.get(
                id=item_id,
                branch=branch,
                stamp_type="order",
                stamp_name=current_stamp
            ).delete()
            # استجابة JSON لو الطلب AJAX
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({"success": True, "deleted_id": item_id})
            else:
                messages.success(request, "🗑️ تم حذف المنتج بنجاح.")
        except StandardRequest.DoesNotExist:
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({"success": False, "message": "❌ المنتج غير موجود."})
            else:
                messages.error(request, "❌ المنتج غير موجود.")
        return redirect("set_standard_request")
    # ❌ حذف كل عناصر الاستمبا الحالية
    if request.method == "POST" and "delete_all" in request.POST:
        StandardRequest.objects.filter(
            branch=branch,
            stamp_type="order",
            stamp_name=current_stamp
        ).delete()
        messages.success(request, f"❌ تم حذف جميع العناصر من '{current_stamp}'.")
        return redirect("set_standard_request")

    # ================== GET / عرض الصفحة ==================
    products = Product.objects.filter(is_available=True)
    categories = Category.objects.all()
    second_categories = SecondCategory.objects.all()

    # العناصر جوه الاستمبا الحالية
    standard_items = StandardRequest.objects.filter(
        branch=branch,
        stamp_type="order",
        stamp_name=current_stamp,
    ).select_related("product__category").order_by(
        "product__category__name",
        "product__name"
    )

    # تجهيز شكل الكمية للعرض
    for item in standard_items:
        if item.product.unit == "kg":
            item.display_quantity = item.default_quantity.quantize(Decimal('0.01'))
        else:
            item.display_quantity = int(item.default_quantity)

    # كل أسماء الاستمبات بتاعت الفرع
    all_stamps = StandardRequest.objects.filter(
        branch=branch,
        stamp_type="order"
    ).values_list("stamp_name", flat=True).distinct()

    return render(request, "orders/set_standard_request.html", {
        "products": products,
        "categories": categories,
        "second_categories": second_categories,
        "requests_today": standard_items,
        "selected_category": selected_category,
        "page_title": "الطلبية القياسية",
        "current_stamp": current_stamp,
        "all_stamps": all_stamps,
    })
#-------------------------------------------------------
@login_required
def control_requests(request):
    profile = getattr(request.user, "userprofile", None)

    # 🚫 لو مش كنترول او ادمن
    if not profile or profile.role not in ["control", "admin"]:
        return render(
            request,
            "orders/no_permission.html",
            {
                "error_message": "🚫 غير مسموح لك بدخول هذه الصفحة. من فضلك تواصل مع مدير النظام لو محتاج صلاحية."
            },
            status=403
        )

    today = timezone.now().date()  # تاريخ اليوم
    branch_id = request.GET.get("branch")
    start_date = request.GET.get("start_date", str(localdate()))
    end_date = request.GET.get("end_date", str(localdate()))
    printed_filter = request.GET.get("printed", "no")

    requests_qs = DailyRequest.objects.filter(is_confirmed=True, created_at__date__range=[start_date, end_date])

    if branch_id:
        requests_qs = requests_qs.filter(branch_id=branch_id)

    if printed_filter == "yes":
        requests_qs = requests_qs.filter(is_printed=True)
    elif printed_filter == "no":
        requests_qs = requests_qs.filter(is_printed=False)

    # Group by (branch, order_number, created_by)
    grouped_requests = {}
    for r in requests_qs.select_related("branch", "product", "created_by").order_by("order_number", "created_at"):
        key = (r.branch, r.order_number, r.created_by)
        grouped_requests.setdefault(key, []).append(r)

    branches = Branch.objects.all()

    return render(request, "orders/control_requests.html", {
        "today": today,
        "grouped_requests": grouped_requests,
        "branches": branches,
        "selected_start": start_date,
        "selected_end": end_date,
        "selected_branch": branch_id,
        "printed_filter": printed_filter,
    })
#-------------------------sockets-----------------------
@login_required
def control_requests_data(request):
    """ترجع HTML الطلبات فقط لتحديث الصفحة عبر AJAX"""
    today = timezone.now().date()
    branch_id = request.GET.get("branch")
    start_date = request.GET.get("start_date", str(localdate()))
    end_date = request.GET.get("end_date", str(localdate()))
    printed_filter = request.GET.get("printed", "no")

    requests_qs = DailyRequest.objects.filter(is_confirmed=True, created_at__date__range=[start_date, end_date])

    if branch_id:
        requests_qs = requests_qs.filter(branch_id=branch_id)

    if printed_filter == "yes":
        requests_qs = requests_qs.filter(is_printed=True)
    elif printed_filter == "no":
        requests_qs = requests_qs.filter(is_printed=False)

    grouped_requests = {}
    for r in requests_qs.select_related("branch", "product", "created_by").order_by("order_number", "created_at"):
        key = (r.branch, r.order_number, r.created_by)
        grouped_requests.setdefault(key, []).append(r)

    html = render_to_string("orders/_requests_list.html", {
        "grouped_requests": grouped_requests,
        "today": today,
    }, request=request)

    return JsonResponse({"html": html})
#-------------------------------------------------------
@require_POST
@login_required
def mark_printed(request, order_number):
    requests = DailyRequest.objects.filter(order_number=order_number)
    if not requests.exists():
        return JsonResponse({"status": "not_found"}, status=404)

    # ✅ تحديث الحالة
    requests.update(is_printed=True, printed_at=timezone.now())

    # ✅ إرسال تحديث للسوكيت
    layer = get_channel_layer()
    async_to_sync(layer.group_send)(
        "control_updates",
        {
            "type": "control_update",
            "action": "printed",
            "message": f"طلبية رقم {order_number} تم تعليمها مطبوعة ✅",
            "order_number": order_number,
        }
    )

    return JsonResponse({"status": "ok"})
#------------------------------------------------------
@login_required
def branch_requests(request):
    profile = getattr(request.user, "userprofile", None)

    # 🚫 لو مش فرع
    if not profile or profile.role not in ["branch"]:
        return render(
            request,
            "orders/no_permission.html",
            {
                "error_message": "🚫 غير مسموح لك بدخول هذه الصفحة."
            },
            status=403
        )

    branch = profile.branch
    today = localdate()

    start_date = request.GET.get("start_date", str(today))
    end_date = request.GET.get("end_date", str(today))
    printed_filter = request.GET.get("printed", "no")

    # ✅ الطلبات المؤكدة الخاصة بالفرع الحالي فقط
    requests_qs = DailyRequest.objects.filter(
        is_confirmed=True,
        branch=branch,
        created_at__date__range=[start_date, end_date]
    )

    if printed_filter == "yes":
        requests_qs = requests_qs.filter(is_printed=True)
    elif printed_filter == "no":
        requests_qs = requests_qs.filter(is_printed=False)

    # ✅ Group by (branch, order_number, created_by)
    grouped_requests = {}
    for r in requests_qs.select_related("product", "created_by").order_by("order_number", "created_at"):
        key = (branch, r.order_number, r.created_by)
        grouped_requests.setdefault(key, []).append(r)

    return render(request, "orders/branch_requests.html", {
        "today": today,
        "grouped_requests": grouped_requests,
        "branches": [branch],          # mirror للكنترول بس للفرع الحالي
        "selected_branch": branch.id,  # يتعلم في الـ select
        "selected_start": start_date,
        "selected_end": end_date,
        "printed_filter": printed_filter,
        "branch": branch,
    })
#-----------------------------------------------------
@login_required
@role_required(["admin"])
def import_products(request):
    if request.method == "POST" and request.FILES.get("excel_file"):
        excel_file = request.FILES["excel_file"]
        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            # ✅ الأعمدة المطلوبة
            expected_headers = [
                "name", "price", "category_name", "second_category_name", "unit", "Is Show"
            ]
            headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]

            required_headers = ["name", "price", "category_name", "second_category_name"]
            if any(h not in headers for h in required_headers):
                messages.error(
                    request,
                    "❌ ملف Excel غير صحيح، يجب أن يحتوي على الأعمدة التالية على الأقل:\n"
                    "name, price, category_name, second_category_name"
                )
                return redirect("import_products")

            header_index = {h: headers.index(h) for h in headers if h in expected_headers}
            count = 0
            hidden_count = 0
            visible_count = 0

            for row in sheet.iter_rows(min_row=2, values_only=True):
                name = row[header_index["name"]] if "name" in header_index else None
                price = row[header_index["price"]] if "price" in header_index else None
                category_name = row[header_index["category_name"]] if "category_name" in header_index else None
                second_category_name = row[header_index["second_category_name"]] if "second_category_name" in header_index else None
                unit_value = row[header_index["unit"]] if "unit" in header_index else None
                is_show_value = row[header_index["Is Show"]] if "Is Show" in header_index else None

                if not name:
                    continue

                # 🔹 السعر
                try:
                    price_value = Decimal(str(price)) if price is not None else Decimal("0.0")
                except (InvalidOperation, TypeError, ValueError):
                    price_value = Decimal("0.0")

                # 🔹 الكاتيجوري الرئيسي
                category = None
                if category_name:
                    category, _ = Category.objects.get_or_create(name=str(category_name).strip())

                # 🔹 الكاتيجوري الفرعي
                second_category = None
                if second_category_name and category:
                    second_category, _ = SecondCategory.objects.get_or_create(
                        name=str(second_category_name).strip(),
                        main_category=category
                    )

                # 🔹 الوحدة
                unit_value = str(unit_value).strip().lower() if unit_value else "piece"
                if unit_value not in ["piece", "kg"]:
                    unit_value = "piece"

                # 🔹 Is Show (عمود الإكسيل)
                is_shwo_clean = None
                is_available = True  # الافتراضي

                if isinstance(is_show_value, str):
                    is_show_value = is_show_value.strip().lower()

                if is_show_value in [True, "true", "yes", "1"]:
                    is_shwo_clean = True
                    is_available = False   # ⬅️ العكس
                    hidden_count += 1
                elif is_show_value in [False, "false", "no", "0"]:
                    is_shwo_clean = False
                    is_available = True    # ⬅️ العكس
                    visible_count += 1
                else:
                    is_shwo_clean = None
                    is_available = True
                    visible_count += 1

                # 🔹 إنشاء أو تحديث المنتج
                product, created = Product.objects.get_or_create(
                    name=str(name).strip(),
                    defaults={
                        "price": price_value,
                        "category": category,
                        "second_category": second_category,
                        "unit": unit_value,
                        "is_available": is_available,
                        "is_shwo": is_shwo_clean,
                    }
                )

                if not created:
                    product.price = price_value
                    product.category = category
                    product.second_category = second_category
                    product.unit = unit_value
                    product.is_available = is_available
                    product.is_shwo = is_shwo_clean
                    product.save()

                count += 1

            messages.success(
                request,
                f"✅ تم استيراد أو تحديث {count} منتج.\n"
                f"📦 المعروضة: {visible_count}, 🚫 المخفية: {hidden_count}"
            )
            return redirect("import_products")

        except Exception as e:
            messages.error(request, f"⚠️ حدث خطأ أثناء قراءة الملف: {e}")
            return redirect("import_products")

    # 📄 GET → عرض الصفحة
    return render(request, "orders/import_products.html")
#-----------------------------------------------------
@login_required
@role_required(["control", "admin","production"])
def set_production_items(request):
    query = request.GET.get("q", "").strip()
    selected_cat = request.GET.get("category", "").strip()
    # current_cat = request.GET.get("current_cat", "").strip()
    # 👇 خليه ياخد من POST أولاً، أو من GET لو مش موجود
    current_cat = request.POST.get("current_cat", request.GET.get("current_cat", "")).strip()

    # الأقسام كلها متاحة للفلترة (اللي فوق)
    categories = Category.objects.all().order_by("name")

    # المنتجات المتاحة للإضافة
    products_qs = Product.objects.filter(is_available=True).select_related("category")
    if selected_cat:
        products_qs = products_qs.filter(category_id=selected_cat)
    if query:
        products_qs = products_qs.filter(name__icontains=query)

    # العناصر الحالية في جدول الإنتاج
    current_items = ProductionTemplate.objects.select_related("product", "product__category").order_by(
        "product__category__name", "product__name"
    )
    if current_cat:
        current_items = current_items.filter(product__category_id=current_cat)

    # 🔥 الأقسام اللي فيها منتجات مضافة فقط (عشان فلتر الجدول)
    current_item_categories = Category.objects.filter(
        products__in=Product.objects.filter(production_templates__isnull=False)
    ).distinct().order_by("name")


    current_product_ids = set(current_items.values_list("product_id", flat=True))

    if request.method == "POST":
        q = request.POST.get("q", "")
        cat = request.POST.get("category", "")
        current_cat = request.POST.get("current_cat", current_cat)
        redirect_url = f"{reverse('set_production_items')}?q={q}&category={cat}&current_cat={current_cat}"

        if "add_product" in request.POST:
            pid = request.POST.get("product_id")
            try:
                p = Product.objects.get(id=pid)
                ProductionTemplate.objects.get_or_create(product=p, defaults={"is_active": True})
                messages.success(request, f"✅ تم إضافة {p.name} لقائمة الإنتاج.")
            except Product.DoesNotExist:
                messages.error(request, "❌ المنتج غير موجود.")
            return redirect(redirect_url)

        elif "toggle_item" in request.POST:
            tid = request.POST.get("toggle_item")
            try:
                t = ProductionTemplate.objects.get(id=tid)
                t.is_active = not t.is_active
                t.save()
                messages.success(request, f"🔁 تم تغيير حالة {t.product.name} إلى {'مفعل' if t.is_active else 'موقوف'}.")
            except ProductionTemplate.DoesNotExist:
                pass
            return redirect(redirect_url)

        elif "delete_item" in request.POST:
            tid = request.POST.get("delete_item")
            ProductionTemplate.objects.filter(id=tid).delete()
            messages.success(request, "🗑️ تم حذف المنتج من قائمة الإنتاج.")
            return redirect(redirect_url)

        elif "delete_all" in request.POST:
            ProductionTemplate.objects.all().delete()
            messages.success(request, "🧹 تم حذف جميع المنتجات من قائمة الإنتاج.")
            return redirect("set_production_items")

        elif "delete_by_category" in request.POST:
            if current_cat:
                ProductionTemplate.objects.filter(product__category_id=current_cat).delete()
                messages.success(request, "🧹 تم حذف المنتجات الخاصة بهذا القسم فقط.")
            else:
                messages.warning(request, "⚠️ لم يتم اختيار قسم حالي من الفلتر.")
            return redirect(redirect_url)
    return render(request, "orders/set_production_items.html", {
        "products": products_qs,
        "current_items": current_items,
        "query": query,
        "categories": categories,  # للفلاتر العليا (اختيار منتجات)
        "current_item_categories": current_item_categories,  # للفلاتر أسفل الجدول
        "selected_cat": selected_cat,
        "current_cat": current_cat,
        "current_product_ids": current_product_ids,
    })
#-----------------------------------------------------
@login_required
@role_required(["branch"])
def add_production_request(request):
    """
    صفحة الفرع: نموذج يومي جاهز يحتوي على المنتجات التي يحددها الكنترول.
    يدعم الفلترة حسب القسم، وكل قسم يتم تأكيده بشكل مستقل.
    """
    profile = getattr(request.user, "userprofile", None)
    branch = profile.branch if profile else None
    if not branch:
        return render(request, "orders/no_permission.html", {
            "error_message": "🚫 لا يوجد فرع مربوط بحسابك."
        }, status=403)

    today = localdate()

    # الأقسام (للـ dropdown)
    categories = Category.objects.filter(id__in=ProductionTemplate.objects.filter(is_active=True).values_list("product__category_id", flat=True)).order_by("name").distinct()

    # فلترة حسب القسم
    selected_cat = request.GET.get("category", "")
    templates = ProductionTemplate.objects.filter(is_active=True).select_related("product", "product__category")

    if selected_cat:
        templates = templates.filter(product__category_id=selected_cat)

    templates = templates.order_by("product__category__name", "product__name")

    # لو مفيش عناصر محددة من الكنترول
    if not templates.exists():
        messages.warning(request, "⚠️ لا توجد منتجات محددة من الكنترول بعد أو لا توجد منتجات في هذا القسم.")
        return render(request, "orders/add_production_request.html", {
            "items": [],
            "today": today,
            "branch": branch,
            "categories": categories,
            "selected_cat": selected_cat,
            "already_confirmed": False,
        })

    # هل الفرع أكد القسم ده النهارده؟
    confirmed_products = ProductionRequest.objects.filter(
        branch=branch, date=today, confirmed=True
    ).values_list("product_id", flat=True)

    existing = ProductionRequest.objects.filter(branch=branch, date=today).select_related("product")
    existing_map = {pr.product_id: pr for pr in existing}

    # POST
    if request.method == "POST":
        # لو فيه قسم محدد، استخدمه في الرابط بعد الحفظ
        redirect_url = f"{reverse('add_production_request')}?category={selected_cat}"

        # تأكد إن الفرع ما أكّدش القسم ده بالكامل قبل كده
        already_confirmed_section = all(t.product_id in confirmed_products for t in templates)
        if already_confirmed_section:
            messages.error(request, "✅ تم تأكيد هذا القسم مسبقًا. لا يمكن التعديل.")
            return redirect(redirect_url)

        saved = 0
        for t in templates:
            pid = t.product_id
            key = f"quantities[{pid}]"
            raw = (request.POST.get(key) or "").strip()
            q = to_decimal_safe(raw, places=2)
            if q < 0:
                q = Decimal("0.00")

            if not unit_allows_fraction(t.product.unit):
                q = q.to_integral_value(rounding=ROUND_HALF_UP)

            obj, _ = ProductionRequest.objects.get_or_create(
                branch=branch, product=t.product, date=today,
                defaults={"quantity": Decimal("0.00"), "created_by": request.user}
            )
            obj.quantity = q
            obj.created_by = request.user
            obj.save()
            saved += 1

        # تأكيد القسم فقط
        if "confirm" in request.POST:
            now_ = timezone.now()
            ProductionRequest.objects.filter(
                branch=branch, date=today, product__in=[t.product for t in templates]
            ).update(confirmed=True, confirmed_at=now_)
            # إشعار الكنترول
            layer = get_channel_layer()
            async_to_sync(layer.group_send)(
                "control_updates",
                {
                    "type": "control_update",
                    "action": "production_confirmed",
                    "message": f"✅ فرع {branch.name} أكد قسم {templates.first().product.category.name} لليوم.",
                }
            )
            messages.success(request, f"✅ تم تأكيد قسم {templates.first().product.category.name} ({saved} صف).")
            return redirect(redirect_url)

        messages.success(request, f"💾 تم حفظ الكميات ({saved} صف) في قسم {templates.first().product.category.name}.")
        return redirect(redirect_url)

    # تجهيز عناصر العرض
    items = []
    for t in templates:
        cur_qty = Decimal("0.00")
        if t.product_id in existing_map:
            cur_qty = to_decimal_safe(existing_map[t.product_id].quantity, places=2)
            if not unit_allows_fraction(t.product.unit):
                cur_qty = cur_qty.to_integral_value()
        items.append({
            "product": t.product,
            "unit": t.product.get_unit_display(),
            "quantity": cur_qty,
            "is_confirmed": t.product_id in confirmed_products
        })

    # هل كل القسم مؤكد؟
    already_confirmed_section = all(i["is_confirmed"] for i in items)

    return render(request, "orders/add_production_request.html", {
        "items": items,
        "today": today,
        "branch": branch,
        "categories": categories,
        "selected_cat": selected_cat,
        "already_confirmed": already_confirmed_section,
    })

#-----------------------------------------------------
@login_required
@role_required(["control", "admin","production"])
def production_overview(request):
    """
    الكنترول: عرض مجمع لطلبات إنتاج اليوم مع فلترة حسب التاريخ، الفرع أو القسم.
    """
    date_raw = request.GET.get("date")
    branch_filter = request.GET.get("branch", "").strip()
    category_filter = request.GET.get("category", "").strip()  # ✅ جديد
    if "hide_zero" in request.GET:
        hide_zero = request.GET.get("hide_zero") == "1"
    else:
        hide_zero = True

    try:
        the_date = datetime.strptime(date_raw, "%Y-%m-%d").date() if date_raw else localdate()
    except ValueError:
        the_date = localdate()

    # الفروع
    branches = list(Branch.objects.all().order_by("name"))
    if branch_filter:
        branches = [b for b in branches if str(b.id) == branch_filter]

    # الأقسام
    categories = Category.objects.filter(
            products__production_templates__is_active=True
        ).distinct().order_by("name")

    # الطلبات
    pr_qs = ProductionRequest.objects.filter(date=the_date).select_related("product", "branch", "product__category")

    # المنتجات الفعالة من قالب الكنترول
    templates = ProductionTemplate.objects.filter(is_active=True).select_related("product", "product__category").order_by(
        "product__category__name", "product__name"
    )

    # ✅ فلترة المنتجات حسب القسم لو المستخدم اختار قسم معين
    if category_filter:
        templates = templates.filter(product__category_id=category_filter)

    products = [t.product for t in templates]

    # خريطة كميات
    cell = {}
    for pr in pr_qs:
        cell[(pr.product_id, pr.branch_id)] = to_decimal_safe(pr.quantity, places=2)

    rows = []
    grand_total = Decimal("0.00")
    for p in products:
        per_branch = []
        row_total = Decimal("0.00")
        for b in Branch.objects.all().order_by("name"):
            if branch_filter and str(b.id) != branch_filter:
                continue
            q = cell.get((p.id, b.id), Decimal("0.00"))
            if not unit_allows_fraction(p.unit):
                q = q.to_integral_value()
            per_branch.append(q)
            row_total += Decimal(str(q))
        if hide_zero and all(q == 0 for q in per_branch):
            continue
        grand_total += row_total
        rows.append({
            "product": p,
            "unit": p.get_unit_display(),
            "per_branch": per_branch,
            "total": row_total
        })

    # اسم الفرع لو محدد
    branch_name = None
    if branch_filter:
        try:
            branch_name = Branch.objects.get(id=branch_filter).name
        except Branch.DoesNotExist:
            branch_name = None
    # 🔍 فحص حالة كل فرع (هل أكّد كل الأقسام المطلوبة؟)
    branch_status = []
    templates_all = ProductionTemplate.objects.filter(is_active=True).values_list("product_id", flat=True)

    for b in Branch.objects.all().order_by("name"):
        total_required = templates_all.count()
        confirmed_count = ProductionRequest.objects.filter(
            branch=b, date=the_date, confirmed=True, product_id__in=templates_all
        ).count()

        # ✅ لو عدد المنتجات المؤكدة = كل المنتجات المطلوبة
        is_done = (confirmed_count == total_required and total_required > 0)

        branch_status.append({
            "branch": b,
            "done": is_done,
            "confirmed": confirmed_count,
            "total": total_required
        })

    return render(request, "orders/production_overview.html", {
        "date": the_date.isoformat(),
        "branches": Branch.objects.all().order_by("name"),
        "categories": categories,  # ✅ جديد
        "rows": rows,
        "grand_total": grand_total,
        "branch_filter": branch_filter,
        "branch_name": branch_name,
        "category_filter": category_filter,  # ✅ جديد
        "branch_status": branch_status,
        "hide_zero": hide_zero
    })
#-----------------------------------------------------
@login_required
@role_required(["control", "admin","production"])
def export_production_excel(request):
    # 🧹 تنظيف البيانات القادمة من الرابط
    date_raw = (request.GET.get("date") or "").strip()
    branch_filter = (request.GET.get("branch") or "").strip()
    category_filter = (request.GET.get("category") or "").strip()
    hide_zero = (request.GET.get("hide_zero") or "1").strip() == "1"

    # 📅 معالجة التاريخ
    try:
        the_date = datetime.strptime(date_raw, "%Y-%m-%d").date() if date_raw else localdate()
    except ValueError:
        the_date = localdate()

    # 🏬 الفروع (فلترة حسب الاختيار)
    branches = list(Branch.objects.all().order_by("name"))
    if branch_filter:
        branches = [b for b in branches if str(b.id) == branch_filter]

    # 🧩 المنتجات الفعالة من قالب الكنترول (ومفلترة بالقسم إن وُجد)
    templates = ProductionTemplate.objects.filter(is_active=True).select_related("product", "product__category")
    if category_filter:
        templates = templates.filter(product__category_id=category_filter)
    templates = templates.order_by("product__category__name", "product__name")
    products = [t.product for t in templates]

    # 🗂️ الطلبات الفعلية حسب التاريخ فقط
    pr_qs = ProductionRequest.objects.filter(date=the_date).select_related("product", "branch")
    if branch_filter:
        pr_qs = pr_qs.filter(branch_id=branch_filter)
    if category_filter:
        pr_qs = pr_qs.filter(product__category_id=category_filter)

    # 🧮 خريطة كميات
    cell = {}
    for pr in pr_qs:
        cell[(pr.product_id, pr.branch_id)] = to_decimal_safe(pr.quantity, places=2)

    # 🧾 إنشاء ملف Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Production {the_date.isoformat()}"

    # ✳️ أول صف توثيقي: التاريخ + الفرع + القسم (بدون حالة الأصفار)
    branch_name = ""
    if branch_filter:
        try:
            branch_name = Branch.objects.get(id=branch_filter).name
        except Branch.DoesNotExist:
            branch_name = ""
    else:
        branch_name = "كل الفروع"

    category_name = ""
    if category_filter:
        try:
            category_name = Category.objects.get(id=category_filter).name
        except Category.DoesNotExist:
            category_name = ""
    else:
        category_name = "كل الأقسام"

    # info_text = f"📅 التاريخ: {the_date.isoformat()} | 🏬 {branch_name} | 📂 {category_name}"
    info_text = f"📅 التاريخ: {the_date.strftime('%Y/%m/%d')} | 🏬 {branch_name} | 📂 {category_name}"

    # 🧩 صف التوثيق
    ws.append([info_text])
    total_cols = len(["المنتج", "الوحدة"] + [b.name for b in branches] + ["الإجمالي"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

    ws["A1"].font = Font(bold=True, size=12, color="000000")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # لون رمادي فاتح

    # 🧩 صف العناوين
    headers = ["المنتج", "الوحدة"] + [b.name for b in branches] + ["الإجمالي"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

    # 🎨 تهيئة رؤوس الجدول
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = align_center
        c.border = thin_border

    # 🧩 تعبئة البيانات
    for p in products:
        per_branch = []
        total = Decimal("0.00")
        for b in branches:
            q = cell.get((p.id, b.id), Decimal("0.00"))
            if not unit_allows_fraction(p.unit):
                q = q.to_integral_value()
            per_branch.append(q)
            total += Decimal(str(q))

        if hide_zero and all(q == 0 for q in per_branch):
            continue

        ws.append([p.name, p.get_unit_display()] + per_branch + [total])

    # ⚙️ تنسيق الأعمدة (تجاوز الصف المدموج الأول)
    for col_cells in ws.iter_cols(min_row=2):
        first_cell = col_cells[0]
        col_letter = first_cell.column_letter
        max_len = 0
        for c in col_cells:
            c.border = thin_border
            c.alignment = align_center
            if c.value:
                max_len = max(max_len, len(str(c.value)))
        ws.column_dimensions[col_letter].width = max_len + 2

    # 📄 إعداد ملف الرد
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"production_{the_date.isoformat()}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

#-----------------------------------------------------
