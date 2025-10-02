from datetime import datetime
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db import transaction
from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.timezone import now

from .models import (
    Applicant, ApplicantHistory, DeletedApplicant, AcceptedApplicant, Queue, STATUS_CHOICES
)
from .forms import (
    ApplicantCreateForm, ApplicantEditFormHRHelp, ApplicantEditFormHR,
    ExperienceFormSet, AcceptedFollowUpForm
)
from .utils import is_hr, is_hr_help,is_hr_or_hr_help,is_admin_or_hr_or_hr_help,is_admin_or_hr,is_admin

from django.contrib.auth.decorators import login_required
from django.shortcuts import render

@login_required
def hr_dashboard(request):
    return render(request, "hr/hr_dashboard.html")

@login_required
def hr_help_dashboard(request):
    return render(request, "hr/hr_help_dashboard.html")


# ------------------------------------------------------------------------------
# HELPER: Ø³Ø¬Ù„ Ø­Ø±ÙƒØ©/ØªØ§Ø±ÙŠØ®
# ------------------------------------------------------------------------------
def _log_history(applicant, user, action, changes=None):
    ApplicantHistory.objects.create(
        applicant=applicant,
        action=action,
        updated_by=user,
        changes=changes or ""
    )

# Ù…Ù‚Ø§Ø±Ù†Ø© Ø¨Ø³ÙŠØ·Ø© Ø¨ÙŠÙ† Ù‚ÙŠÙ… Ù‚Ø¯ÙŠÙ…Ø©/Ø¬Ø¯ÙŠØ¯Ø© Ø¹Ø´Ø§Ù† Ù†Ø³Ø¬Ù„ "Ø¥ÙŠÙ‡ Ø§Ù„Ù„ÙŠ Ø§ØªØºÙŠØ±"
def _diff_changes(instance, old_data: dict, fields):
    diffs = []
    for f in fields:
        old = old_data.get(f, None)
        new = getattr(instance, f, None)
        if str(old) != str(new):
            diffs.append(f"{f}: '{old}' -> '{new}'")
    return "; ".join(diffs)

# ------------------------------------------------------------------------------
# 1) Ø¥Ù†Ø´Ø§Ø¡ Ø·Ù„Ø¨ Ø¬Ø¯ÙŠØ¯ (HR_HELP ÙÙ‚Ø·)
# ------------------------------------------------------------------------------
@login_required
@user_passes_test(is_admin_or_hr_or_hr_help)
@transaction.atomic
def applicant_create(request):
    if request.method == "POST":
        form = ApplicantCreateForm(request.POST, request.FILES)
        formset = ExperienceFormSet(request.POST, prefix="exp")
        if form.is_valid() and formset.is_valid():
            # Ø­Ù…Ø§ÙŠØ© ØªÙƒØ±Ø§Ø± Ø§Ù„Ø±Ù‚Ù… Ø§Ù„Ù‚ÙˆÙ…ÙŠ Ø¨Ø±Ø³Ø§Ù„Ø© Ø£ÙˆØ¶Ø­
            nid = form.cleaned_data.get("national_id")
            if Applicant.objects.filter(national_id=nid).exists():
                messages.error(request, "Ù‡Ø°Ø§ Ø§Ù„Ø±Ù‚Ù… Ø§Ù„Ù‚ÙˆÙ…ÙŠ Ù…Ø³Ø¬Ù‘Ù„ Ø¨Ø§Ù„ÙØ¹Ù„ â€” Ù„Ø§ ÙŠÙ…ÙƒÙ† ØªÙƒØ±Ø§Ø± Ø§Ù„Ø·Ù„Ø¨.")
                return redirect("hr:applicant_search_or_create")

            applicant = form.save(commit=False)
            applicant.created_by = request.user
            applicant.last_updated_by = request.user
            applicant.save()

            formset.instance = applicant
            formset.save()

            _log_history(applicant, request.user, "create", changes="Ø¥Ù†Ø´Ø§Ø¡ Ø·Ù„Ø¨ Ø¬Ø¯ÙŠØ¯")

            messages.success(request, f"ØªÙ… Ø­ÙØ¸ Ø§Ù„Ø·Ù„Ø¨ #{applicant.order_number} Ø¨Ù†Ø¬Ø§Ø­.")
            return redirect("hr:applicant_detail", order_number=applicant.order_number)
        else:
            messages.error(request, "Ø¨Ø±Ø¬Ø§Ø¡ Ù…Ø±Ø§Ø¬Ø¹Ø© Ø§Ù„Ù…Ø¯Ø®Ù„Ø§Øª.")
    else:
        form = ApplicantCreateForm()
        formset = ExperienceFormSet(prefix="exp")

    # âœ… Ø¬Ø±ÙˆØ¨Ø§Øª Ø§Ù„ÙÙŠÙ„Ø¯Ø² Ø¹Ø´Ø§Ù† Ù†Ø³ØªØ®Ø¯Ù…Ù‡Ø§ ÙÙŠ Ø§Ù„Ù€ template Ø¨Ø¯Ù„ split
    fields_basic = [
        "national_id", "full_name", "phone", "marital_status",
        "nationality", "gender", "religion", "military_status",
        "email", "relative_name", "relative_phone",
        "is_smoker", "vehicle_ownership", "photo"
    ]

    fields_edu = [
        "edu_degree", "grad_year", "edu_institution",
        "specialization", "postgrad_study", "edu_grade"
    ]

    fields_job = [
        "job_applied", "job_code", "prev_applied",
        "has_relatives_in_company", "relatives_in_company",
        "has_relatives_in_competitors", "relatives_in_competitors",
        "has_health_issues", "health_issues_details"
    ]

    return render(request, "hr/applicant_create.html", {
        "form": form,
        "formset": formset,
        "page_title": "ØªØ³Ø¬ÙŠÙ„ Ø·Ù„Ø¨ Ø¬Ø¯ÙŠØ¯",
        "fields_basic": fields_basic,
        "fields_edu": fields_edu,
        "fields_job": fields_job,
    })
#---------------------------------------------------------------------------------
# Ø´Ø§Ø´Ø© Ø¨Ø­Ø« Ø³Ø±ÙŠØ¹Ø© Ù‚Ø¨Ù„ Ø§Ù„Ø¥Ù†Ø´Ø§Ø¡ (ÙŠØ¯Ø®Ù„ Ø§Ù„Ø±Ù‚Ù… Ø§Ù„Ù‚ÙˆÙ…ÙŠ)
@login_required
@user_passes_test(is_admin_or_hr_or_hr_help)
def applicant_search_or_create(request):
    nid = request.GET.get("national_id", "").strip()
    found = None
    if nid:
        found = Applicant.objects.filter(national_id=nid).first()
        if found:
            messages.warning(request, f"Ø§Ù„Ù…ØªÙ‚Ø¯Ù… Ø¨Ø±Ù‚Ù… Ù‚ÙˆÙ…ÙŠ {nid} Ù„Ø¯ÙŠÙ‡ Ø·Ù„Ø¨ Ø³Ø§Ø¨Ù‚ #{found.order_number}.")
    return render(request, "hr/applicant_search_or_create.html", {
        "found": found,
        "query_nid": nid,
    })

# ------------------------------------------------------------------------------
# 2) Ø¹Ø±Ø¶ Ù‚Ø§Ø¦Ù…Ø© Ø§Ù„Ø·Ù„Ø¨Ø§Øª + ÙÙ„ØªØ±Ø© (HR ÙŠØ±Ù‰ Ø§Ù„ÙƒÙ„ / HR_HELP ÙŠØ±Ù‰ Ù…Ø§ Ø£Ù†Ø´Ø£Ù‡)
# ------------------------------------------------------------------------------
from datetime import datetime
from django.utils.timezone import now
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from .models import Applicant, STATUS_CHOICES
from .utils import is_hr, is_hr_help

@login_required
def applicant_list(request):
    qs = Applicant.objects.all().order_by("order_number")
    role_hr = is_hr(request.user)
    role_help = is_hr_help(request.user)

    if role_help and not role_hr:
        qs = qs.filter(created_by=request.user)  # Ø§Ù„Ù…ÙˆØ¸Ù Ø§Ù„Ø¹Ø§Ø¯ÙŠ ÙŠØ´ÙˆÙ Ø·Ù„Ø¨Ø§ØªÙ‡ ÙÙ‚Ø·

    # ÙÙ„Ø§ØªØ±
    today_date = now().date()
    today_str = today_date.strftime("%Y-%m-%d")

    date_from = request.GET.get("from") or today_str
    date_to = request.GET.get("to") or today_str
    status = request.GET.get("status")
    q = request.GET.get("q")

    # ÙÙ„ØªØ±Ø© Ø¨Ø§Ù„ØªÙˆØ§Ø±ÙŠØ®
    try:
        dt_from = datetime.strptime(date_from, "%Y-%m-%d").date()
        if dt_from <= today_date:
            qs = qs.filter(created_at__date__gte=dt_from)
    except ValueError:
        pass

    try:
        dt_to = datetime.strptime(date_to, "%Y-%m-%d").date()
        if dt_to <= today_date:
            qs = qs.filter(created_at__date__lte=dt_to)
    except ValueError:
        pass

    if status and status in dict(STATUS_CHOICES):
        qs = qs.filter(status=status)

    if q:
        qs = (
            qs.filter(full_name__icontains=q)
            | qs.filter(national_id__icontains=q)
            | qs.filter(phone__icontains=q)
        )

    return render(request, "hr/applicant_list.html", {
        "applicants": qs,
        "status_choices": STATUS_CHOICES,
        "filters": {
        "from": date_from or "",
        "to": date_to or "",
        "status": status or "",
        "q": q or ""},
        "today": today_str,
    })

# ------------------------------------------------------------------------------
# 3) ØªÙØ§ØµÙŠÙ„ Ø·Ù„Ø¨
# ------------------------------------------------------------------------------
@login_required
def applicant_detail(request, order_number):
    applicant = get_object_or_404(Applicant, order_number=order_number)
    # Ø§Ù„Ù…ÙˆØ¸Ù Ø§Ù„Ø¹Ø§Ø¯ÙŠ Ù„Ø§ ÙŠØ±Ù‰ Ø¥Ù„Ø§ Ù…Ø§ Ø£Ù†Ø´Ø£Ù‡
    if is_hr_help(request.user) and not is_hr(request.user):
        if applicant.created_by_id != request.user.id:
            raise Http404("ØºÙŠØ± Ù…Ø³Ù…ÙˆØ­ Ù„Ùƒ Ø¨Ø§Ù„Ø§Ø·Ù„Ø§Ø¹ Ø¹Ù„Ù‰ Ù‡Ø°Ø§ Ø§Ù„Ø·Ù„Ø¨.")
    return render(request, "hr/applicant_detail.html", {
        "applicant": applicant,
    })

# ------------------------------------------------------------------------------
# 4) ØªØ¹Ø¯ÙŠÙ„ Ø·Ù„Ø¨ (HR_HELP ÙŠØ¹Ø¯Ù„ Ø·Ù„Ø¨Ø§ØªÙ‡ØŒ HR ÙŠØ¹Ø¯Ù„ Ø£ÙŠ Ø·Ù„Ø¨)
# ------------------------------------------------------------------------------
@login_required
@transaction.atomic
def applicant_edit(request, order_number):
    applicant = get_object_or_404(Applicant, order_number=order_number)
    role_hr = is_hr(request.user)
    role_help = is_hr_help(request.user)
    role_admin=is_admin(request.user)
    # ØµÙ„Ø§Ø­ÙŠØ§Øª Ø§Ù„ÙˆØµÙˆÙ„
    if role_help and not (role_hr or role_admin) and applicant.created_by_id != request.user.id:
        raise Http404("ØºÙŠØ± Ù…Ø³Ù…ÙˆØ­ Ù„Ùƒ Ø¨ØªØ¹Ø¯ÙŠÙ„ Ù‡Ø°Ø§ Ø§Ù„Ø·Ù„Ø¨.")

    # Ø§Ø®ØªÙŠØ§Ø± Ø§Ù„ÙÙˆØ±Ù… Ø­Ø³Ø¨ Ø§Ù„Ø¯ÙˆØ±
    FormClass = ApplicantEditFormHR if role_hr else ApplicantEditFormHRHelp

    if request.method == "POST":
        form = FormClass(request.POST, request.FILES, instance=applicant)
        formset = ExperienceFormSet(request.POST, prefix="exp", instance=applicant)
        if form.is_valid() and formset.is_valid():
            # Ø®Ø²Ù† Ù‚ÙŠÙ… Ù‚Ø¯ÙŠÙ…Ø© Ø¹Ø´Ø§Ù† Ù†Ø³Ø¬Ù„ Ø§Ù„ÙØ±Ù‚
            tracked_fields = [f.name for f in Applicant._meta.fields if f.name not in ("created_at", "updated_at")]
            old_data = {f: getattr(applicant, f) for f in tracked_fields}

            obj = form.save(commit=False)
            obj.last_updated_by = request.user
            obj.save()
            formset.save()

            # Ø³Ø¬Ù„ Ø§Ù„Ø§Ø®ØªÙ„Ø§ÙØ§Øª
            changes = _diff_changes(obj, old_data, tracked_fields)
            _log_history(obj, request.user, "update", changes=changes or "ØªØ¹Ø¯ÙŠÙ„ Ø¨ÙŠØ§Ù†Ø§Øª")

            messages.success(request, "ØªÙ… Ø­ÙØ¸ Ø§Ù„ØªØ¹Ø¯ÙŠÙ„Ø§Øª.")
            return redirect("hr:applicant_detail", order_number=order_number)
        else:
            messages.error(request, "Ø¨Ø±Ø¬Ø§Ø¡ Ù…Ø±Ø§Ø¬Ø¹Ø© Ø§Ù„Ù…Ø¯Ø®Ù„Ø§Øª.")
    else:
        form = FormClass(instance=applicant)
        formset = ExperienceFormSet(prefix="exp", instance=applicant)

    return render(request, "hr/applicant_edit.html", {
        "form": form,
        "formset": formset,
        "applicant": applicant,
        "fields_basic": [
            "national_id", "full_name", "phone", "marital_status", "nationality",
            "gender", "religion", "military_status", "email", "relative_name",
            "relative_phone", "is_smoker", "vehicle_ownership", "photo"
        ],
        "fields_edu": [
            "edu_degree", "grad_year", "edu_institution", "specialization",
            "postgrad_study", "edu_grade"
        ],
        "fields_job": [
            "job_applied", "job_code", "prev_applied", "has_relatives_in_company",
            "relatives_in_company", "has_relatives_in_competitors",
            "relatives_in_competitors", "has_health_issues",
            "health_issues_details", "status"
        ],
    })

# ------------------------------------------------------------------------------
# 5) Ø­Ø°Ù Ø·Ù„Ø¨ (HR ÙÙ‚Ø·) + ØªØ³Ø¬ÙŠÙ„ ÙÙŠ DeletedApplicant + History
# ------------------------------------------------------------------------------
@login_required
@user_passes_test(is_admin_or_hr)
@transaction.atomic
def applicant_delete(request, order_number):
    applicant = get_object_or_404(Applicant, order_number=order_number)

    if request.method == "POST":
        # Ø³Ø¬Ù„ Ù†Ø³Ø®Ø© Ù„Ù„Ù…Ø­Ø°ÙˆÙØ§Øª
        DeletedApplicant.objects.create(
            original_order_number=applicant.order_number,
            full_name=applicant.full_name,
            national_id=applicant.national_id,
            phone=applicant.phone,
            email=applicant.email,
            deleted_by=request.user,
        )
        _log_history(applicant, request.user, "delete", changes="Ø­Ø°Ù Ø§Ù„Ø·Ù„Ø¨")

        applicant.delete()
        messages.success(request, "ØªÙ… Ø­Ø°Ù Ø§Ù„Ø·Ù„Ø¨ØŒ ÙˆÙŠÙ…ÙƒÙ†Ùƒ Ù…Ø±Ø§Ø¬Ø¹ØªÙ‡ ÙÙŠ ØµÙØ­Ø© Ø§Ù„Ù…Ø­Ø°ÙˆÙØ§Øª.")
        return redirect("hr:applicant_list")

    return render(request, "hr/applicant_delete_confirm.html", {
        "applicant": applicant,
    })

# ------------------------------------------------------------------------------
# 6) Ø§ØªØ®Ø§Ø° Ù‚Ø±Ø§Ø± HR (Ù‚Ø¨ÙˆÙ„/Ø±ÙØ¶/Ø§Ø­ØªÙŠØ§Ø·ÙŠ) + Ø¥Ù†Ø´Ø§Ø¡ Ù†Ø³Ø®Ø© Ø¨Ø§Ù„Ù…Ù‚Ø¨ÙˆÙ„ÙŠÙ†
# ------------------------------------------------------------------------------
@login_required
@user_passes_test(is_admin_or_hr)
@transaction.atomic
def applicant_decision(request, order_number):
    applicant = get_object_or_404(Applicant, order_number=order_number)
    if request.method == "POST":
        decision = request.POST.get("decision")  # accepted/rejected/reserve
        if decision not in dict(STATUS_CHOICES):
            messages.error(request, "Ù‚Ø±Ø§Ø± ØºÙŠØ± ØµØ§Ù„Ø­.")
            return redirect("hr:applicant_detail", order_number=order_number)

        applicant.status = decision
        applicant.decision_by = request.user
        applicant.decision_at = timezone.now()
        applicant.last_updated_by = request.user
        applicant.save()

        # Ù„Ùˆ Ù…Ù‚Ø¨ÙˆÙ„ â€” Ø£Ù†Ø´Ø¦ Ø³Ø¬Ù„ AcceptedApplicant Ù„Ùˆ Ù…Ø´ Ù…ÙˆØ¬ÙˆØ¯
        if decision == "accepted":
            AcceptedApplicant.objects.get_or_create(applicant=applicant)

        _log_history(applicant, request.user, "decision", changes=f"Ù‚Ø±Ø§Ø±: {decision}")
        messages.success(request, "ØªÙ… Ø­ÙØ¸ Ø§Ù„Ù‚Ø±Ø§Ø±.")
        return redirect("hr:applicant_detail", order_number=order_number)

    return render(request, "hr/applicant_decision.html", {
        "applicant": applicant,
        "status_choices": STATUS_CHOICES,
    })

# ------------------------------------------------------------------------------
# 7) Ù‚Ø§Ø¦Ù…Ø© Ø§Ù„Ù…Ù‚Ø¨ÙˆÙ„ÙŠÙ† (ÙÙ„ØªØ± ØªØ§Ø±ÙŠØ®)
# ------------------------------------------------------------------------------
from datetime import datetime, date
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from .models import AcceptedApplicant

@login_required
def accepted_list(request):
    qs = AcceptedApplicant.objects.select_related("applicant").all().order_by("applicant__order_number")
    date_from = request.GET.get("from")
    date_to = request.GET.get("to")

    today = date.today()

    # ğŸŸ¢ Ù„Ùˆ Ù…ÙÙŠØ´ Ù‚ÙŠÙ… Ù…Ù† Ø§Ù„ÙÙˆØ±Ù… Ø®Ù„ÙŠÙ‡Ù… Ø¨ØªØ§Ø±ÙŠØ® Ø§Ù„Ù†Ù‡Ø§Ø±Ø¯Ù‡
    if not date_from:
        date_from = today.strftime("%Y-%m-%d")
    if not date_to:
        date_to = today.strftime("%Y-%m-%d")

    dt_from, dt_to = None, None
    try:
        dt_from = datetime.strptime(date_from, "%Y-%m-%d").date()
    except ValueError:
        dt_from = today

    try:
        dt_to = datetime.strptime(date_to, "%Y-%m-%d").date()
    except ValueError:
        dt_to = today

    # âœ… Ù„Ùˆ from Ø£ÙƒØ¨Ø± Ù…Ù† to Ø®Ù„ÙŠÙ‡Ù… Ù…ØªØ³Ø§ÙˆÙŠÙŠÙ†
    if dt_from > dt_to:
        dt_from = dt_to
        date_from = date_to

    qs = qs.filter(applicant__decision_at__date__gte=dt_from,
                   applicant__decision_at__date__lte=dt_to)

    return render(request, "hr/accepted_list.html", {
        "accepted": qs,
        "filters": {"from": date_from, "to": date_to},
        "today": today.strftime("%Y-%m-%d"),
    })


# ------------------------------------------------------------------------------
# 8) Ù‚Ø§Ø¦Ù…Ø© Ø§Ù„Ù…Ø­Ø°ÙˆÙØ§Øª (ÙÙ„ØªØ± ØªØ§Ø±ÙŠØ®)
# ------------------------------------------------------------------------------
from datetime import datetime, date
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render
from .models import DeletedApplicant
from .utils import is_admin_or_hr

@login_required
@user_passes_test(is_admin_or_hr)
def deleted_list(request):
    qs = DeletedApplicant.objects.all().order_by("-snapshot_at")
    date_from = request.GET.get("from")
    date_to = request.GET.get("to")

    today = date.today()

    # ğŸŸ¢ Ù„Ùˆ Ù…ÙÙŠØ´ Ù‚ÙŠÙ… Ù…Ù† Ø§Ù„ÙÙˆØ±Ù… Ø®Ù„ÙŠÙ‡Ù… Ø¨ØªØ§Ø±ÙŠØ® Ø§Ù„Ù†Ù‡Ø§Ø±Ø¯Ù‡
    if not date_from:
        date_from = today.strftime("%Y-%m-%d")
    if not date_to:
        date_to = today.strftime("%Y-%m-%d")

    dt_from, dt_to = None, None
    try:
        dt_from = datetime.strptime(date_from, "%Y-%m-%d").date()
    except ValueError:
        dt_from = today

    try:
        dt_to = datetime.strptime(date_to, "%Y-%m-%d").date()
    except ValueError:
        dt_to = today

    # âœ… Ù„Ùˆ from Ø£ÙƒØ¨Ø± Ù…Ù† to Ø®Ù„ÙŠÙ‡Ù… Ù…ØªØ³Ø§ÙˆÙŠÙŠÙ†
    if dt_from > dt_to:
        dt_from = dt_to
        date_from = date_to

    qs = qs.filter(snapshot_at__date__gte=dt_from,
                   snapshot_at__date__lte=dt_to)

    return render(request, "hr/deleted_list.html", {
        "deleted": qs,
        "filters": {"from": date_from, "to": date_to},
        "today": today.strftime("%Y-%m-%d"),
    })

# ------------------------------------------------------------------------------
# 9) Ø´Ø§Ø´Ø© Ø§Ù„Ø¯ÙˆØ± (HR Ùˆ HR_HELP ÙŠØ´ÙˆÙÙˆØ§ / HR ÙŠØ¶ØºØ· Ø§Ù„ØªØ§Ù„ÙŠ)
# ------------------------------------------------------------------------------
def _get_next_pending(current_order_number=None):
    qs = Applicant.objects.filter(status="pending").order_by("order_number")
    if current_order_number:
        return qs.filter(order_number__gt=current_order_number).first() or qs.first()
    return qs.first()

@login_required
def queue_view(request):
    queue, _ = Queue.objects.get_or_create(id=1)
    # Ø£ÙˆÙ„ Ù…Ø±Ø©: Ù„Ùˆ Ù…ÙÙŠØ´ Ø­Ø¯ Ù…Ø­Ø¯Ø¯ ÙÙŠ Ø§Ù„Ø¯ÙˆØ± â€” Ø­Ø¯Ø¯ Ø£ÙˆÙ„ pending
    if not queue.current_applicant:
        queue.current_applicant = _get_next_pending()
        queue.save()

    # Ø§Ù„Ù…ÙˆØ¸Ù Ø§Ù„Ø¹Ø§Ø¯ÙŠ ÙŠØ´ÙˆÙ Ø¨Ø³ (Ø´Ø§Ø´Ø© Ø®Ø¶Ø±Ø§Ø¡/ØªÙ†Ø¨ÙŠÙ‡ Ø¨Ø±Ù‚Ù… Ø§Ù„Ø·Ù„Ø¨ Ø§Ù„Ø­Ø§Ù„ÙŠ)
    return render(request, "hr/queue.html", {
        "current": queue.current_applicant,
        "can_next": is_admin_or_hr(request.user),  # Ø²Ø±Ø§Ø± Ø§Ù„ØªØ§Ù„ÙŠ ÙŠØ¸Ù‡Ø± Ù„Ù€ HR ÙÙ‚Ø·
    })
#-------------------------------------------------------------------------------------------------------
@login_required
@user_passes_test(is_admin_or_hr)
def queue_next(request):
    queue, _ = Queue.objects.get_or_create(id=1)
    current = queue.current_applicant
    next_app = _get_next_pending(current.order_number if current else None)
    queue.current_applicant = next_app
    queue.save()
    if next_app:
        messages.success(request, f"Ø§Ù„ØªØ§Ù„ÙŠ: Ø·Ù„Ø¨ #{next_app.order_number} - {next_app.full_name}")
    else:
        messages.info(request, "Ù„Ø§ ÙŠÙˆØ¬Ø¯ Ø·Ù„Ø¨Ø§Øª Ù‚ÙŠØ¯ Ø§Ù„Ø§Ù†ØªØ¸Ø§Ø±.")
    return redirect("hr:queue_view")

# ------------------------------------------------------------------------------
# 10) ØªØµØ¯ÙŠØ± Ø¥ÙƒØ³Ù„ (Ø·Ù„Ø¨ ÙˆØ§Ø­Ø¯ / Ù…Ø¬Ù…Ù‘Ø¹)
# ------------------------------------------------------------------------------
import openpyxl
from openpyxl.utils import get_column_letter

def _worksheet_autofit(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                value = str(cell.value) if cell.value is not None else ""
                if len(value) > max_length:
                    max_length = len(value)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2
#-------------------------------------------------------------------------------------------------------
@login_required
def export_applicant_excel(request, order_number):
    applicant = get_object_or_404(Applicant, order_number=order_number)
    # ØµÙ„Ø§Ø­ÙŠØ§Øª Ø§Ù„Ø§Ø·Ù„Ø§Ø¹: HR ÙŠØ´ÙˆÙ Ø§Ù„ÙƒÙ„ / HR_HELP ÙŠØ´ÙˆÙ Ø·Ù„Ø¨Ø§ØªÙ‡
    if is_hr_help(request.user) and not is_hr(request.user):
        if applicant.created_by_id != request.user.id:
            raise Http404("ØºÙŠØ± Ù…Ø³Ù…ÙˆØ­ Ù„Ùƒ.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Ø·Ù„Ø¨ #{applicant.order_number}"

    rows = [
        ("Ø±Ù‚Ù… Ø§Ù„Ø·Ù„Ø¨", applicant.order_number),
        ("Ø§Ù„Ø§Ø³Ù…", applicant.full_name),
        ("Ø§Ù„Ø±Ù‚Ù… Ø§Ù„Ù‚ÙˆÙ…ÙŠ", applicant.national_id),
        ("Ø§Ù„Ù‡Ø§ØªÙ", applicant.phone),
        ("Ø§Ù„Ø­Ø§Ù„Ø© Ø§Ù„Ø§Ø¬ØªÙ…Ø§Ø¹ÙŠØ©", applicant.get_marital_status_display()),
        ("Ø§Ù„Ø¬Ù†Ø³ÙŠØ©", applicant.get_nationality_display()),
        ("Ø§Ù„Ù†ÙˆØ¹", applicant.get_gender_display()),
        ("Ø§Ù„Ø¯ÙŠØ§Ù†Ø©", applicant.get_religion_display()),
        ("Ø§Ù„Ù…ÙˆÙ‚Ù Ù…Ù† Ø§Ù„ØªØ¬Ù†ÙŠØ¯", applicant.get_military_status_display()),
        ("Ø§Ù„Ø¨Ø±ÙŠØ¯", applicant.email or ""),
        ("Ø§Ø³Ù… Ø§Ù„Ù‚Ø±ÙŠØ¨", applicant.relative_name),
        ("Ù‡Ø§ØªÙ Ø§Ù„Ù‚Ø±ÙŠØ¨", applicant.relative_phone),
        ("Ù…Ø¯Ø®Ù†", "Ù†Ø¹Ù…" if applicant.is_smoker else "Ù„Ø§"),
        ("ÙˆØ³ÙŠÙ„Ø©/Ø³ÙŠØ§Ø±Ø©", applicant.get_vehicle_ownership_display()),

        ("Ø§Ù„Ù…Ø¤Ù‡Ù„", applicant.get_edu_degree_display() if applicant.edu_degree else ""),
        ("Ø³Ù†Ø© Ø§Ù„ØªØ®Ø±Ø¬", applicant.grad_year or ""),
        ("Ø¬Ù‡Ø© Ø§Ù„Ø­ØµÙˆÙ„", applicant.edu_institution or ""),
        ("Ø§Ù„ØªØ®ØµØµ", applicant.specialization or ""),
        ("Ø¯Ø±Ø§Ø³Ø§Øª Ø¹Ù„ÙŠØ§", applicant.postgrad_study or ""),
        ("Ø§Ù„ØªÙ‚Ø¯ÙŠØ±", applicant.get_edu_grade_display() if applicant.edu_grade else ""),

        ("Ø§Ù„ÙˆØ¸ÙŠÙØ© Ø§Ù„Ù…ØªÙ‚Ø¯Ù… Ù„Ù‡Ø§", applicant.get_job_applied_display()),
        ("ÙƒÙˆØ¯ Ø§Ù„ÙˆØ¸ÙŠÙØ©", applicant.job_code or ""),
        ("ØªØ§Ø±ÙŠØ® Ø§Ù„ØªÙ‚Ø¯ÙŠÙ…", applicant.submitted_at.strftime("%Y-%m-%d %H:%M:%S")),
        ("Ø³Ø¨Ù‚ Ø§Ù„ØªÙ‚Ø¯ÙŠÙ… Ø¨Ø§Ù„Ø´Ø±ÙƒØ©", "Ù†Ø¹Ù…" if applicant.prev_applied else "Ù„Ø§"),
        ("Ø£Ù‚Ø§Ø±Ø¨ Ø¨Ø§Ù„Ø´Ø±ÙƒØ©", "Ù†Ø¹Ù…" if applicant.has_relatives_in_company else "Ù„Ø§"),
        ("ØªÙØ§ØµÙŠÙ„ Ø£Ù‚Ø§Ø±Ø¨ Ø¨Ø§Ù„Ø´Ø±ÙƒØ©", applicant.relatives_in_company or ""),
        ("Ø£Ù‚Ø§Ø±Ø¨/Ø£ØµØ¯Ù‚Ø§Ø¡ Ø¨Ø´Ø±ÙƒØ§Øª Ù…Ù†Ø§ÙØ³Ø©", "Ù†Ø¹Ù…" if applicant.has_relatives_in_competitors else "Ù„Ø§"),
        ("ØªÙØ§ØµÙŠÙ„ Ø£Ù‚Ø§Ø±Ø¨ Ø¨Ø´Ø±ÙƒØ§Øª Ù…Ù†Ø§ÙØ³Ø©", applicant.relatives_in_competitors or ""),
        ("Ù…Ø´Ø§ÙƒÙ„ ØµØ­ÙŠØ©", "Ù†Ø¹Ù…" if applicant.has_health_issues else "Ù„Ø§"),
        ("ØªÙØ§ØµÙŠÙ„ Ø§Ù„Ù…Ø´Ø§ÙƒÙ„ Ø§Ù„ØµØ­ÙŠØ©", applicant.health_issues_details or ""),

        ("Ø§Ù„Ø­Ø§Ù„Ø© Ø§Ù„Ø­Ø§Ù„ÙŠØ©", applicant.get_status_display()),
        ("Ø£ÙÙ†Ø´Ø¦ Ø¨ÙˆØ§Ø³Ø·Ø©", applicant.created_by.get_full_name() if applicant.created_by else ""),
        ("ØªØ§Ø±ÙŠØ® Ø§Ù„Ø¥Ù†Ø´Ø§Ø¡", applicant.created_at.strftime("%Y-%m-%d %H:%M:%S")),
        ("Ø¢Ø®Ø± ØªØ¹Ø¯ÙŠÙ„ Ø¨ÙˆØ§Ø³Ø·Ø©", applicant.last_updated_by.get_full_name() if applicant.last_updated_by else ""),
        ("Ø¢Ø®Ø± ØªØ¹Ø¯ÙŠÙ„", applicant.updated_at.strftime("%Y-%m-%d %H:%M:%S")),
        ("Ø§Ù„Ù‚Ø±Ø§Ø± Ø¨ÙˆØ§Ø³Ø·Ø©", applicant.decision_by.get_full_name() if applicant.decision_by else ""),
        ("ØªØ§Ø±ÙŠØ® Ø§Ù„Ù‚Ø±Ø§Ø±", applicant.decision_at.strftime("%Y-%m-%d %H:%M:%S") if applicant.decision_at else ""),
    ]
    ws.append(["Ø§Ù„Ø­Ù‚Ù„", "Ø§Ù„Ù‚ÙŠÙ…Ø©"])
    for r in rows:
        ws.append(r)

    # Ø´ÙŠØª Ù„Ù„Ø®Ø¨Ø±Ø§Øª Ø§Ù„Ø³Ø§Ø¨Ù‚Ø©
    ws2 = wb.create_sheet("Ø§Ù„Ø®Ø¨Ø±Ø§Øª Ø§Ù„Ø³Ø§Ø¨Ù‚Ø©")
    ws2.append(["Ø¬Ù‡Ø© Ø§Ù„Ø¹Ù…Ù„", "Ø§Ù„ÙˆØ¸ÙŠÙØ©", "Ø§Ù„Ø³Ù†ÙˆØ§Øª", "Ø§Ù„Ø±Ø§ØªØ¨", "Ø³Ø¨Ø¨ ØªØ±Ùƒ Ø§Ù„Ø¹Ù…Ù„"])
    for exp in applicant.experiences.all():
        ws2.append([exp.employer, exp.job_title, exp.years, float(exp.salary), exp.reason_for_leaving])

    _worksheet_autofit(ws)
    _worksheet_autofit(ws2)

    filename = f"applicant_{applicant.order_number}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
#-------------------------------------------------------------------------------------------------------
@login_required
def export_applicants_excel(request):
    # HR: Ø§Ù„ÙƒÙ„ â€” HR_HELP: Ø¨ØªØ§Ø¹Ù‡ ÙÙ‚Ø·
    qs = Applicant.objects.all().order_by("order_number")
    if is_hr_help(request.user) and not is_hr(request.user):
        qs = qs.filter(created_by=request.user)

    # ÙÙ„Ø§ØªØ± Ø§Ø®ØªÙŠØ§Ø±ÙŠØ©: ØªØ§Ø±ÙŠØ® Ù…Ù†/Ø¥Ù„Ù‰
    date_from = request.GET.get("from")
    date_to = request.GET.get("to")
    if date_from:
        try:
            dt = datetime.strptime(date_from, "%Y-%m-%d")
            qs = qs.filter(created_at__date__gte=dt.date())
        except ValueError:
            pass
    if date_to:
        try:
            dt = datetime.strptime(date_to, "%Y-%m-%d")
            qs = qs.filter(created_at__date__lte=dt.date())
        except ValueError:
            pass

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ø§Ù„Ø·Ù„Ø¨Ø§Øª"
    ws.append([
        "Ø±Ù‚Ù… Ø§Ù„Ø·Ù„Ø¨", "Ø§Ù„Ø§Ø³Ù…", "Ø§Ù„Ø±Ù‚Ù… Ø§Ù„Ù‚ÙˆÙ…ÙŠ", "Ø§Ù„Ù‡Ø§ØªÙ", "Ø§Ù„Ø­Ø§Ù„Ø©", "ØªØ§Ø±ÙŠØ® Ø§Ù„ØªÙ‚Ø¯ÙŠÙ…", "Ø£ÙÙ†Ø´Ø¦ Ø¨ÙˆØ§Ø³Ø·Ø©"
    ])
    for a in qs:
        ws.append([
            a.order_number, a.full_name, a.national_id, a.phone,
            a.get_status_display(),
            a.submitted_at.strftime("%Y-%m-%d %H:%M:%S"),
            a.created_by.get_full_name() if a.created_by else "",
        ])

    _worksheet_autofit(ws)
    filename = "applicants.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
